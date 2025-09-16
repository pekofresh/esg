# -*- coding: utf-8 -*-
"""
Created on Tue Sep 17 11:29:19 2024

@author: A00008106
"""

import os
import sys
import json
import time
import logging
from dataclasses import dataclass, field
from functools import wraps
from typing import Any, Sequence, Tuple, Optional, Union, Callable, Dict

import numpy as np
import pandas as pd

pd.set_option("display.max_columns", 500)
pd.set_option("display.width", 1000)

import pickle
import math
# import matplotlib.pyplot as plt
from statsmodels.distributions.empirical_distribution import ECDF
from datetime import datetime

# import blpapi
# import pdblp
from pulp import *
import lib_utils_AP as lib_utils

import openpyxl
from openpyxl import Workbook
import re

from pyspark.sql.types import StringType


# ---------------------------------------------------------------------------
# Configuration and logging
# ---------------------------------------------------------------------------


@dataclass
class SmaConfig:
    """Central configuration for lib_sma_AP paths and table names."""

    ids_holdings_table: str = field(
        default_factory=lambda: os.getenv(
            "DREMIO_IDS_HOLDINGS_TABLE", lib_utils.dictDremio["dfIDSHoldings"]
        )
    )
    ids_fund_table: str = field(
        default_factory=lambda: os.getenv(
            "DREMIO_IDS_FUND_TABLE", lib_utils.dictDremio["dfIDSFund"]
        )
    )
    country_dim_table: str = field(
        default_factory=lambda: os.getenv(
            "DREMIO_COUNTRY_DIM_TABLE", lib_utils.dictDremio["dfCountryDim"]
        )
    )
    portfolio_dim_table: str = field(
        default_factory=lambda: os.getenv(
            "DREMIO_PORTFOLIO_DIM_TABLE", lib_utils.dictDremio["dfPortfolioDim"]
        )
    )
    target_funds_cache_parquet: str = field(
        default_factory=lambda: os.getenv(
            "TARGET_FUNDS_CACHE_PARQUET",
            "dbfs:/Volumes/zenith_dev_gold/esg_impact_analysis/cache_volume/dfExt.parquet",
        )
    )
    target_funds_cache_pickle: str = field(
        default_factory=lambda: os.getenv(
            "TARGET_FUNDS_CACHE_PICKLE",
            os.path.join(lib_utils.sTargetFundsFolder, "dfExt.pkl"),
        )
    )
    countries_cache_path: str = field(
        default_factory=lambda: os.getenv(
            "COUNTRY_CACHE_PATH",
            "dbfs:/Volumes/zenith_dev_gold/esg_impact_analysis/cache_volume/dfCountries.parquet",
        )
    )
    cedar_cache_path: str = field(
        default_factory=lambda: os.getenv(
            "CEDAR_CACHE_PATH",
            "dbfs:/Volumes/zenith_dev_gold/esg_impact_analysis/cache_volume/dfCEDAR.parquet",
        )
    )
    refinitiv_data_dir: str = field(
        default_factory=lambda: os.getenv("REFINITIV_DATA_DIR", "")
    )


CONFIG = SmaConfig()


class JsonFormatter(logging.Formatter):
    """Minimal JSON log formatter."""

    def format(self, record: logging.LogRecord) -> str:  # type: ignore[override]
        log_record = {
            "time": self.formatTime(record, self.datefmt),
            "level": record.levelname,
            "name": record.name,
            "message": record.getMessage(),
        }
        return json.dumps(log_record)


logger = logging.getLogger(__name__)
if not logger.handlers:
    handler = logging.StreamHandler()
    handler.setFormatter(JsonFormatter())
    logger.addHandler(handler)
logger.setLevel(logging.INFO)


def log_timing(func):
    """Decorator to log the execution time of functions."""

    @wraps(func)
    def wrapper(*args, **kwargs):
        start = time.perf_counter()
        result = func(*args, **kwargs)
        duration = time.perf_counter() - start
        logger.info("%s completed in %.2fs", func.__name__, duration)
        return result

    return wrapper


def _build_in_clause(values: Sequence[Any]) -> Tuple[str, list[Any]]:
    """Return a parametrised SQL IN clause and corresponding parameters."""

    if not values:
        raise ValueError("Sequence cannot be empty")
    placeholders = ", ".join(["?"] * len(values))
    return f"({placeholders})", list(values)


def _read_refinitiv_file(file_path: str) -> pd.DataFrame:
    """Read a single Refinitiv CSV file and filter valid holdings."""
    df = pd.read_csv(file_path)
    return df.loc[df.Allocation_Type == "FLHLD"]


# load master data
@log_timing
def _load_refinitiv(
    refresh: bool,
    spark=None,
    data_dir: str | None = None,
    cache_path: str | None = None,
    exclude_files: set[str] | None = None,
) -> pd.DataFrame:
    """Load Refinitiv target fund holdings.

    Parameters
    ----------
    refresh : bool
        Rebuild cache from raw CSV files when True, otherwise load cached data.
    spark : SparkSession, optional
        Active Spark session used to read/write Parquet caches on DBFS.
    data_dir : str, optional
        Directory containing the Refinitiv CSV files. Defaults to the
        ``REFINITIV_DATA_DIR`` environment variable.
    cache_path : str, optional
        Location of the cached DataFrame. Defaults to a parquet file on DBFS if
        ``spark`` is provided, otherwise to a local pickle file in
        ``lib_utils.sTargetFundsFolder``.
    exclude_files : set[str], optional
        File names to skip while scanning ``data_dir``.
    """

    if data_dir is None:
        data_dir = CONFIG.refinitiv_data_dir
        if not data_dir:
            raise EnvironmentError(
                "REFINITIV_DATA_DIR environment variable or data_dir parameter must be set"
            )

    if exclude_files is None:
        exclude_files = lib_utils.REFINITIV_EXCLUDE_FILES

    if cache_path is None:
        if spark:
            cache_path = CONFIG.target_funds_cache_parquet
        else:
            cache_path = CONFIG.target_funds_cache_pickle

    logger.info("Loading target fund holdings...")

    if refresh:
        files = [
            f for f in os.listdir(data_dir) if f not in exclude_files
        ]
        if not files:
            raise FileNotFoundError(f"No Refinitiv files found in {data_dir}")

        frames = []
        for fname in files:
            file_path = os.path.join(data_dir, fname)
            logger.debug("Processing %s", fname)
            frames.append(_read_refinitiv_file(file_path))

        df_out = pd.concat(frames, ignore_index=True)
        df_out.Allocation_Percentage = df_out.Allocation_Percentage.astype(float)
        df_out = df_out.drop_duplicates()

        df_out = lib_utils.apply_column_mappings(
            df_out,
            rename_map=lib_utils.REFINITIV_RENAME_COLUMNS,
            drop_cols=lib_utils.REFINITIV_DROP_COLUMNS,
            fill_values=lib_utils.REFINITIV_PRE_MERGE_CONSTANTS,
        )

        df_out = pd.merge(
            df_out,
            lib_utils.dfSecurities,
            how="left",
            left_on="id_isin",
            right_on="ID_ISIN",
            validate="m:1",
        )

        df_out = lib_utils.apply_column_mappings(
            df_out,
            copy_map=lib_utils.REFINITIV_POST_MERGE_COPY,
            fill_values=lib_utils.REFINITIV_POST_MERGE_CONSTANTS,
        )

        df_out["weight"] = df_out["weight"] / 100

        if spark:
            save_cache(df_out, cache_path, spark)
        else:
            os.makedirs(os.path.dirname(cache_path), exist_ok=True)
            df_out.to_pickle(cache_path)

    else:
        if spark:
            df_out = load_cache(cache_path, spark)
        else:
            df_out = pd.read_pickle(cache_path)

    return df_out

logger = logging.getLogger(__name__)
@log_timing
# -----------------------------------------------------------------------------
# SQL helpers
# -----------------------------------------------------------------------------
def _build_in_clause(values: Sequence[Any], prefix: str = "param") -> tuple[str, dict[str, Any]]:
    """
    Build a safe SQL IN clause with parameter placeholders.

    Args:
        values: Values to include in the IN clause.
        prefix: Prefix for parameter names.

    Returns:
        A tuple (clause, params), where:
          - clause is a SQL string like "( :param0, :param1 )"
          - params is a dict mapping placeholders to values
    """
    if not values:
        raise ValueError("IN clause values must not be empty.")

    clause_parts = []
    params: dict[str, Any] = {}
    for i, val in enumerate(values):
        key = f"{prefix}{i}"
        clause_parts.append(f":{key}")
        params[key] = val

    return f"({', '.join(clause_parts)})", params


def build_ids_portfolio_sql(
    ids_holdings_table: str,
    ids_fund_table: str,
    dates: Sequence[str],
    fund_ids: Sequence[int],
) -> tuple[str, dict[str, Any]]:
    """Return SQL query and bound params for IDS portfolio holdings."""
    date_clause, date_params = _build_in_clause(dates, prefix="date")
    fund_clause, fund_params = _build_in_clause(fund_ids, prefix="fund")

    sql = f"""
        SELECT a.edm_fund_id AS portfolio_id,
               :benchmark_id AS benchmark_id,
               b.fund_name AS portfolio_name,
               a.instrument_identifier AS id_isin,
               a.instrument_weight_in_fund AS weight,
               CAST(a.validity_date AS DATE) AS effective_date
        FROM {ids_holdings_table} a
        LEFT JOIN (
            SELECT DISTINCT edm_fund_id, fund_name
            FROM {ids_fund_table}
            WHERE edm_fund_id IN {fund_clause}
        ) b
          ON a.edm_fund_id = b.edm_fund_id
        WHERE CAST(a.validity_date AS DATE) IN {date_clause}
          AND a.edm_fund_id IN {fund_clause}
    """

    params = {**date_params, **fund_params}
    return sql, params


# -----------------------------------------------------------------------------
# DataFrame enrichments & normalizers
# -----------------------------------------------------------------------------
def normalize_company_ids(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure Bloomberg company IDs are numeric floats."""
    for col in ["ID_BB_COMPANY", "ID_BB_PARENT_CO", "ID_BB_ULTIMATE_PARENT_CO"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def enrich_with_securities(df: pd.DataFrame, df_securities: pd.DataFrame, join_key: str) -> pd.DataFrame:
    """Join securities reference metadata to holdings."""
    return df.merge(
        df_securities,
        how="left",
        left_on=join_key,
        right_on="ID_ISIN",
        validate="m:1",
    )


def enrich_with_countries(df: pd.DataFrame, df_country: pd.DataFrame) -> pd.DataFrame:
    """Join country reference metadata to holdings."""
    out = df.merge(
        df_country,
        how="left",
        left_on="ULT_PARENT_CNTRY_DOMICILE",
        right_on="country_code",
        validate="m:1",
    )
    out["country_issue_name"] = out.get("country_name")
    return out


def standardize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Standardize common descriptive columns."""
    if "LONG_COMP_NAME" in df:
        df["issuer_long_name"] = df["LONG_COMP_NAME"]
    if "SECURITY_DES" in df:
        df["long_name"] = df["SECURITY_DES"]
    if "INDUSTRY_SECTOR" in df:
        df["gics_sector_description"] = df["INDUSTRY_SECTOR"].str.upper()
    return df


# -----------------------------------------------------------------------------
# Main loader
# -----------------------------------------------------------------------------
def load_ids_portfolio(
    dates: Sequence[str],
    join_key: str,
    conn: Connection,
    ids_holdings_table: str,
    ids_fund_table: str,
    df_securities: pd.DataFrame,
    df_country: pd.DataFrame,
    benchmark_id: int = 100,
    fund_ids: Sequence[int] = (40002847, 40011162, 40002955, 40002956),
) -> pd.DataFrame:
    """
    Load IDS portfolio holdings for the given dates and enrich with securities and country data.

    Args:
        dates: One or more validity dates ("YYYY-MM-DD").
        join_key: Column name in holdings to merge with securities (typically "id_isin").
        conn: SQLAlchemy connection to Dremio/warehouse.
        ids_holdings_table: Fully qualified name of the IDS holdings table.
        ids_fund_table: Fully qualified name of the IDS fund metadata table.
        df_securities: Preloaded securities reference DataFrame.
        df_country: Preloaded country reference DataFrame.
        benchmark_id: Benchmark ID to attach to all rows (default=100).
        fund_ids: Restrict query to these fund IDs.

    Returns:
        A pandas DataFrame with holdings enriched with security and country metadata.

    Raises:
        ValueError: If dates or fund_ids are empty.
    """
    if not dates:
        raise ValueError("Parameter 'dates' must not be empty.")
    if not fund_ids:
        raise ValueError("Parameter 'fund_ids' must not be empty.")

    sql, params = build_ids_portfolio_sql(ids_holdings_table, ids_fund_table, dates, fund_ids)
    params["benchmark_id"] = benchmark_id

    logger.info(
        "Loading IDS portfolio holdings",
        extra={"dates": dates, "fund_ids": fund_ids, "benchmark_id": benchmark_id},
    )

    df = pd.read_sql(sql, conn, params=params)

    # Enrichment pipeline
    df["durable_key_type"] = "xls"
    df["portfolio_type"] = "Fund"

    df = (
        df.pipe(enrich_with_securities, df_securities, join_key)
          .pipe(normalize_company_ids)
          .pipe(standardize_columns)
          .pipe(enrich_with_countries, df_country)
    )

    return df




def save_cache(df_pandas, path, spark, format='parquet'):
    """
    Save DataFrame to DBFS or local filesystem using Parquet (default) or Pickle.
    
    Args:
        df_pandas (pd.DataFrame): DataFrame to save
        path (str): Path to save to (Parquet path or Pickle file path)
        spark (SparkSession): Active Spark session
        format (str): 'parquet' or 'pickle' (default = 'parquet')
    """
    if format.lower() == 'parquet':
        df_spark = spark.createDataFrame(df_pandas)

        # Handle VOID types
        for col_name, dtype in df_spark.dtypes:
            if dtype.lower() == "void":
                logger.warning(
                    "Casting VOID column %s to StringType for Parquet compatibility",
                    col_name,
                )
                df_spark = df_spark.withColumn(
                    col_name, df_spark[col_name].cast(StringType())
                )

        df_spark.write.mode("overwrite").parquet(path)
    
    elif format.lower() == 'pickle':
        # Convert dbfs:/ to /dbfs/ for local file writing in Databricks
        if path.startswith("dbfs:/"):
            path = "/dbfs/" + path[6:]
        os.makedirs(os.path.dirname(path), exist_ok=True)
        df_pandas.to_pickle(path)
    
    else:
        raise ValueError("Invalid format. Supported formats: 'parquet', 'pickle'.")

def load_cache(path, spark, format='parquet'):
    """
    Load DataFrame from DBFS or local filesystem using Parquet (default) or Pickle.

    Args:
        path (str): Path to load from (Parquet path or Pickle file path)
        spark (SparkSession): Active Spark session
        format (str): 'parquet' or 'pickle' (default = 'parquet')
    
    Returns:
        pd.DataFrame
    """
    if format.lower() == 'parquet':
        return spark.read.parquet(path).toPandas()
    
    elif format.lower() == 'pickle':
        if path.startswith("dbfs:/"):
            path = "/dbfs/" + path[6:]
        return pd.read_pickle(path)
    
    else:
        raise ValueError("Invalid format. Supported formats: 'parquet', 'pickle'.")

logger = logging.getLogger(__name__)


def build_country_sql(country_dim_table: str) -> str:
    """Return SQL query to fetch all countries."""
    return f"SELECT * FROM {country_dim_table}"


def load_cntry(
    refresh: bool,
    spark: SparkSession,
    country_dim_table: str,
    cache_path: Optional[str] = None,
) -> pd.DataFrame:
    """
    Load country dimension data, either from cache or by querying Spark.

    Args:
        refresh: If True, query Spark and refresh cache; if False, load from cache.
        spark: SparkSession for executing SQL.
        country_dim_table: Fully qualified table name for country dimension.
        cache_path: Path to cached file (parquet or pickle). Required if refresh=False.

    Returns:
        Pandas DataFrame with country dimension data.
    """
    if not cache_path and not refresh:
        raise ValueError("cache_path is required when refresh=False")

    if refresh:
        sql = build_country_sql(country_dim_table)
        logger.info("Refreshing country data", extra={"table": country_dim_table})
        df = spark.sql(sql).toPandas()
        save_cache(df, cache_path, spark)
    else:
        logger.info("Loading country data from cache", extra={"path": cache_path})
        df = load_cache(cache_path, spark)

    return df

@log_timing
def build_cedar_sql(portfolio_dim_table: str) -> str:
    """Return SQL query to fetch current CEDAR portfolios."""
    return f"""
        SELECT *
        FROM {portfolio_dim_table}
        WHERE is_current = TRUE
    """


def filter_cedar(
    df: pd.DataFrame,
    identifiers: Optional[Sequence[str]] = None,
    column: Optional[str] = None,
) -> pd.DataFrame:
    """
    Filter CEDAR DataFrame by identifier values in a given column
    and restrict to active portfolios.

    Args:
        df: CEDAR portfolio DataFrame.
        identifiers: List of identifier values to search for.
        column: Column to apply identifier filter to.

    Returns:
        Filtered DataFrame.
    """
    if identifiers and column:
        pattern = "|".join(map(str, identifiers))
        mask = df[column].str.contains(pattern, na=False)
        df = df[mask]
    return df[df["portfolio_status"] == "Active"]


def search_cedar(
    refresh: bool,
    spark: SparkSession,
    portfolio_dim_table: str,
    cache_path: str,
    identifiers: Optional[Sequence[str]] = None,
    column: Optional[str] = None,
) -> pd.DataFrame:
    """
    Load and optionally filter CEDAR portfolios.

    Args:
        refresh: If True, refresh cache from Spark; else load from cache.
        spark: SparkSession.
        portfolio_dim_table: Fully qualified CEDAR portfolio dimension table.
        cache_path: Path for cache file (parquet/pickle).
        identifiers: Values to match inside column (optional).
        column: Column name to filter on (optional).

    Returns:
        Filtered pandas DataFrame with selected portfolio fields.
    """
    if refresh:
        sql = build_cedar_sql(portfolio_dim_table)
        logger.info("Refreshing CEDAR portfolios", extra={"table": portfolio_dim_table})
        df = spark.sql(sql).toPandas()
        save_cache(df, cache_path, spark)
    else:
        logger.info("Loading CEDAR portfolios from cache", extra={"path": cache_path})
        df = load_cache(cache_path, spark)

    df = filter_cedar(df, identifiers, column)

    return df[
        [
            "portfolio_id",
            "portfolio_name",
            "portfolio_status",
            "bbg_account",
            "strategy_name_long",
            "esg_categorization",
        ]
    ]
def calc_universe(
    name: str,
    sDate: List[str],
    loader: str,                  # "holdings" or "sql"
    source: str | int,            # holdings index ID or SQL query
    esg_source: Dict | None = None,
    cutoff: Dict | None = None,
    export_path: str | None = None,
) -> pd.DataFrame:
    """
    Generic universe builder with optional ESG enrichment and cutoff/threshold logic.
    Exports both the universe and thresholds sheet (if cutoff is applied).
    """

    # ---- load ----
    if loader == "holdings":
        df = load_holdings("index", source, sDate, position_type=1)
    elif loader == "sql":
        df = pd.read_sql(source, conngenie)
    else:
        raise ValueError("Unknown loader type")
    df["Scope"] = 1

    # ---- enrich ESG ----
    if esg_source:
        df_esg = load_dremio_data(esg_source["table"], esg_source["cols"],
                                  sDateCol=esg_source["key"], iMastering=True)
        df = merge_esg_data_new(df, df_esg, esg_source["cols"], "_ESG", False)

    # ---- cutoff logic ----
    thresholds = None
    if cutoff:
        group_cols = cutoff["group"]
        col = cutoff["col"]

        if cutoff["type"] == "median":
            thresholds = df.groupby(group_cols)[col].median().reset_index()
        elif cutoff["type"] == "quantile":
            thresholds = df.groupby(group_cols)[col].quantile(cutoff["value"]).reset_index()
        else:
            raise ValueError("Unsupported cutoff type")

        thresholds = thresholds.rename(columns={col: f"{col}_TH"})
        df = pd.merge(df, thresholds, on=group_cols, how="left")

        if cutoff["type"] == "quantile":
            df["Restricted"] = df[col] < df[f"{col}_TH"]
        else:  # median cutoff
            df["Restricted"] = df[col] >= df[f"{col}_TH"]

    # ---- export ----
    if export_path:
        sheets = [{"data": df, "sSheetName": name}]
        if thresholds is not None:
            sheets.append({"data": thresholds, "sSheetName": "Thresholds"})
        excel_export(pd.ExcelWriter(export_path, engine="xlsxwriter"), sheets)

    return df



# -----------------------------------------------------------------------------------
# convenience wrappers for each universe
# -----------------------------------------------------------------------------------

def calc_universe_Oktagon(sDate: List[str], iExport=True) -> pd.DataFrame:
    return calc_universe(
        name="Oktagon",
        sDate=sDate,
        loader="holdings",
        source=584,
        esg_source={"table": "dfSRI", "cols": ["sri_rating_final"], "key": "effective_date"},
        cutoff=None,
        export_path=f"Universes/Oktagon_{sDate[0]}.xlsx" if iExport else None,
    )


def calc_universe_Unicredit(sDate: List[str], iExport=True) -> pd.DataFrame:
    return calc_universe(
        name="Unicredit",
        sDate=sDate,
        loader="sql",
        source="select * from SU.dbo.tcSU_GHGint_cutoffs_univ_for_SMA",
        cutoff={"type": "median",
                "group": ["GICSSector", "Region"],
                "col": "msci_carbon_emissions_scope_12_inten"},
        export_path=f"Universes/Unicredit_{sDate[0]}.xlsx" if iExport else None,
    )


def calc_universe_Best_Styles_PSR(sDate: List[str], iExport=True) -> pd.DataFrame:
    return calc_universe(
        name="PSR",
        sDate=sDate,
        loader="sql",
        source="select * from SU.dbo.tcSU_SRI_cutoffs_univ_for_SMA",
        esg_source={"table": "dfSRI", "cols": ["pss_score_final_absolute"], "key": "effective_date"},
        cutoff={"type": "quantile",
                "group": ["GICSSector", "Region"],
                "col": "pss_score_final_absolute",
                "value": 0.2},
        export_path=f"Universes/PSR_{sDate[0]}.xlsx" if iExport else None,
    )

def load_sec_hierarchy(
    refresh: bool,
    spark: SparkSession,
    table: str,
    cache_path: str,
    columns: Optional[list[str]] = None,
) -> pd.DataFrame:
    """
    Load security hierarchy either from Spark (if refresh=True) or from cache.

    Args:
        refresh: If True, reload from Spark and overwrite cache.
        spark: Spark session.
        table: Fully qualified Dremio table name.
        cache_path: Path to parquet/pickle cache.
        columns: Optional list of columns to select instead of '*'.

    Returns:
        Securities DataFrame.
    """
    if refresh:
        col_str = "*" if not columns else ", ".join(columns)
        sql = f"SELECT {col_str} FROM {table}"
        logger.info("Refreshing securities from Spark", extra={"table": table, "cache": cache_path})
        df = spark.sql(sql).toPandas()
        save_cache(df, cache_path, spark)
    else:
        logger.info("Loading securities from cache", extra={"cache": cache_path})
        df = load_cache(cache_path, spark)

    if df.empty:
        logger.warning("Securities DataFrame is empty!", extra={"table": table, "cache": cache_path})

    return df

def load_fund_perf(
    pm_last_name: str,
    date_str: str,
    con,
    dict_dremio: dict,
) -> pd.DataFrame:
    """
    Load portfolio and benchmark performance for a given PM and date.

    Args:
        pm_last_name: Portfolio manager's last name (string).
        date_str: Valuation date in 'YYYY-MM-DD' format.
        con: Database connection (e.g., lib_utils.conndremio).
        dict_dremio: Mapping of table keys to Dremio table names.

    Returns:
        DataFrame with portfolio and benchmark performance.
    """
    valuation_date_key = datetime.strptime(date_str, "%Y-%m-%d").strftime("%Y%m%d")

    sql = f"""
        SELECT b.ret_mtd,
               e.ret_mtd AS ret_mtd_benchmark,
               b.ret_ytd,
               e.ret_ytd AS ret_ytd_benchmark,
               a.portfolio_durable_key,
               a.portfolio_id,
               a.portfolio_name,
               d.benchmark_durable_key,
               d.benchmark_name,
               a.bbg_account,
               a.gidp_reporting_location,
               a.investment_region,
               a.sri_esg,
               a.asset_class,
               a.investment_approach,
               a.market_cap
        FROM   {dict_dremio['dfPortfolioDim']} a
        LEFT JOIN {dict_dremio['dfPortfolioPerf']} b
               ON a.portfolio_durable_key = b.portfolio_durable_key
        LEFT JOIN {dict_dremio['dfPortfolioBenchmarkBridge']} c
               ON a.portfolio_durable_key = c.portfolio_durable_key
        LEFT JOIN {dict_dremio['dfBenchmarkDim']} d
               ON c.benchmark_durable_key = d.benchmark_durable_key
        LEFT JOIN {dict_dremio['dfBenchmarkPerf']} e
               ON c.benchmark_durable_Key = e.benchmark_durable_key
              AND e.portfolio_durable_key = b.portfolio_durable_key
              AND b.valuation_date_key = e.valuation_date_key
              AND e.perf_type_key = 2
              AND e.ret_type_key = 2
        WHERE  a.portfolio_durable_key IN (
                   SELECT portfolio_durable_key
                   FROM   {dict_dremio['dfPortfolioManager']}
                   WHERE  last_name = ?
               )
          AND a.is_current = TRUE
          AND b.valuation_date_key = ?
          AND b.perf_type_key = 6
          AND b.ret_type_key = 2
          AND d.is_current = TRUE
        ORDER BY a.portfolio_name
    """

    logger.info("Loading fund performance", extra={"pm_last_name": pm_last_name, "date": valuation_date_key})
    df_out = pd.read_sql(sql, con, params=[pm_last_name, valuation_date_key])

    if df_out.empty:
        logger.warning("No performance records found", extra={"pm_last_name": pm_last_name, "date": valuation_date_key})

    return df_out

def load_green_bonds(
    con: Any,
    dict_dremio: dict,
    columns: list[str] | None = None,
) -> pd.DataFrame:
    """
    Load and filter green/social/sustainability bonds.

    Args:
        con: Database connection (e.g., lib_utils.conndremio).
        dict_dremio: Mapping of logical names to Dremio table names.
        columns: Optional explicit list of columns to select.

    Returns:
        DataFrame with ISINs flagged as green bonds.
    """
    base_cols = ["ID_ISIN", "GREEN_BOND_LOAN_INDICATOR", "SOCIAL_BOND_IND", "SUSTAINABILITY_BOND_IND"]
    col_str = "*" if columns is None else ", ".join(columns)

    sql = f"SELECT {col_str} FROM {dict_dremio['dfGB']}"
    logger.info("Loading green bonds from Dremio", extra={"table": dict_dremio['dfGB']})

    df = pd.read_sql(sql, con)

    # Ensure only relevant columns
    df = df[base_cols].drop_duplicates(subset=["ID_ISIN"])

    # Keep bonds with at least one positive indicator
    mask = df[["GREEN_BOND_LOAN_INDICATOR", "SOCIAL_BOND_IND", "SUSTAINABILITY_BOND_IND"]].max(axis=1) == 1
    df = df.loc[mask].copy()

    df["Exposure_GB"] = True

    if df.empty:
        logger.warning("No green bonds found in table", extra={"table": dict_dremio['dfGB']})

    return df

# load ESG data
def load_c4f_data(
    filename: str,
    base_path: str,
    conn: Any,
    df_securities: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """
    Load C4F data from Excel, enrich with securities hierarchy, and aggregate.

    Args:
        filename: Excel file name (e.g., 'c4f_input.xlsx').
        base_path: Root working directory for Impact Assessment data.
        conn: Database connection (used if df_securities is None).
        df_securities: Optional preloaded securities DataFrame.
                      If None, load_sec_hierarchy will be called.

    Returns:
        Aggregated DataFrame with one row per company.
    """
    work_folder = Path(base_path) / "15. Investment process" / "Impact Assessment"

    if df_securities is None:
        logger.info("Loading securities hierarchy for C4F merge")
        df_securities = load_sec_hierarchy(conn, refresh=False)

    file_path = work_folder / filename
    logger.info("Loading C4F data", extra={"file": str(file_path)})

    df_c4f = pd.read_excel(file_path)

    df_c4f = pd.merge(
        df_c4f,
        df_securities,
        how="left",
        left_on="isin",
        right_on="ID_ISIN",
        validate="m:1",
    )

    # Keep only relevant columns, deduplicate, aggregate
    df_c4f = (
        df_c4f[["ID_BB_COMPANY", "msappbtiPerBeurosTurnoverAdjusted"]]
        .drop_duplicates()
        .sort_values(["ID_BB_COMPANY", "msappbtiPerBeurosTurnoverAdjusted"])
        .groupby("ID_BB_COMPANY")
        .last()
        .reset_index()
    )

    if df_c4f.empty:
        logger.warning("C4F dataset is empty after processing", extra={"file": str(file_path)})

    return df_c4f

def convert_date(
    date_str: str,
    input_fmt: str,
    output_fmt: str,
) -> str:
    """
    Convert a date string from one format to another.

    Args:
        date_str: Input date string (e.g., "2024-09-16").
        input_fmt: Format of the input string (e.g., "%Y-%m-%d").
        output_fmt: Desired output format (e.g., "%Y%m%d").

    Returns:
        The reformatted date string.

    Raises:
        ValueError: If the input string does not match the input format.

    Example:
        >>> convert_date("2024-09-16", "%Y-%m-%d", "%Y%m%d")
        "20240916"
    """
    try:
        parsed = datetime.strptime(date_str, input_fmt)
        return parsed.strftime(output_fmt)
    except ValueError as e:
        logger.error("Date conversion failed", extra={
            "date_str": date_str,
            "input_fmt": input_fmt,
            "output_fmt": output_fmt
        })
        raise

def categorize_market_cap(
    market_cap: Optional[float],
    small_threshold: float = 2e9,
    large_threshold: float = 10e9,
) -> Union[str, float]:
    """
    Categorize a market capitalization value into size buckets.

    Args:
        market_cap: Market cap value in USD.
        small_threshold: Upper bound for 'small' (default 2e9).
        large_threshold: Lower bound for 'large' (default 10e9).

    Returns:
        'small', 'mid', 'large', or np.nan if input is missing.

    Example:
        >>> categorize_market_cap(1.5e9)
        'small'
        >>> categorize_market_cap(5e9)
        'mid'
        >>> categorize_market_cap(15e9)
        'large'
        >>> categorize_market_cap(None)
        nan
    """
    if pd.isna(market_cap):
        return np.nan
    if market_cap < small_threshold:
        return "small"
    if market_cap < large_threshold:
        return "mid"
    return "large"

# registry of transformers
MASTERING_TRANSFORMERS: Dict[str, Callable[[pd.DataFrame], pd.DataFrame]] = {}

def register_mastering(name: str):
    """Decorator to register a mastering function."""
    def wrapper(func: Callable[[pd.DataFrame], pd.DataFrame]):
        MASTERING_TRANSFORMERS[name] = func
        return func
    return wrapper

# -----------------------------------------------------------------------------
# Dataset-specific transformers
# -----------------------------------------------------------------------------

@register_mastering("dfSusMinExclusions")
def _susmin(df: pd.DataFrame) -> pd.DataFrame:
    return df.loc[df.id_exl_list == 3]


@register_mastering("dfSusMinExclusionsEnh")
def _susmin_enh(df: pd.DataFrame) -> pd.DataFrame:
    return df.loc[df.id_exl_list == 65]


@register_mastering("dfTowardsSustainability")
def _towards(df: pd.DataFrame) -> pd.DataFrame:
    return df.loc[df.id_exl_list.isin([6, 7, 53])]


@register_mastering("dfFebelfin")
def _febelfin(df: pd.DataFrame) -> pd.DataFrame:
    return df.loc[df.id_exl_list == 6]


@register_mastering("dfFirmwide")
def _firmwide(df: pd.DataFrame) -> pd.DataFrame:
    return df.loc[df.id_exl_list == 2]


@register_mastering("dfFH")
def _fh(df: pd.DataFrame) -> pd.DataFrame:
    df = df.loc[df.edition == 2024].copy()
    df["status"] = df["status"].eq("NF")
    df["FH Total Score < 35"] = df["total"] < 35
    return df


@register_mastering("dfSRI")
def _sri(df: pd.DataFrame) -> pd.DataFrame:
    df = df.loc[df.universe_name == "GlobalEquityExEurope"].copy()
    thresholds = {
        "sri_rating_final": [1, 1.5, 2],
        "final_override": [1, 1.5, 2],
    }
    for col, vals in thresholds.items():
        for thr in vals:
            df[f"{col} < {thr}"] = df[col] < thr
    return df


@register_mastering("dfPSR")
def _psr(df: pd.DataFrame) -> pd.DataFrame:
    df = df.loc[df.universe_name == "PSS"].copy()
    for thr in [1, 2]:
        df[f"pss_score_final_relative_postflags < {thr}"] = (
            df["pss_score_final_relative_postflags"] < thr
        )
    return df


@register_mastering("dfMSCI")
def _msci(df: pd.DataFrame) -> pd.DataFrame:
    df["issuer_market_cap_usd"] = pd.to_numeric(df["issuer_market_cap_usd"], errors="coerce")
    mask = df["issuer_market_cap_usd"].notnull()
    df.loc[mask, "market_cap_sector"] = pd.cut(
        df.loc[mask, "issuer_market_cap_usd"],
        bins=[0, 300e6, 2e9, 10e9, 200e9, 10e12],
        labels=["micro cap", "small cap", "mid cap", "large cap", "mega cap"],
        include_lowest=True,
    )
    df["market_cap_sector"] = df["market_cap_sector"].astype(str).replace({"nan": "_OTHER"})
    df["market_cap_category"] = df["issuer_market_cap_usd"].apply(
        lambda x: categorize_market_cap(x) if pd.notnull(x) else None
    )
    return df


@register_mastering("dfKPI")
def _kpi(df: pd.DataFrame) -> pd.DataFrame:
    df["carbon_emissions_evic_scope_123_inten"] = (
        df["carbon_emissions_evic_scope_12_inten"] + df["carbon_emissions_scope_3_tot_evic_inten"]
    )
    df["evic_usd"] = df["evic_eur"].astype(float) * 1.17 * 1_000_000
    df["evic_category"] = df["evic_usd"].apply(
        lambda x: categorize_market_cap(x) if pd.notnull(x) else None
    )
    return df


@register_mastering("dfSus")
def _sus(df: pd.DataFrame) -> pd.DataFrame:
    df["non_zero_sustainable_investment_share_environmental_post_dnsh"] = (
        df["sustainable_investment_share_environmental_post_dnsh"].fillna(0).astype(bool)
    )
    df["non_zero_sustainable_investment_share_social_post_dnsh"] = (
        df["sustainable_investment_share_social_post_dnsh"].fillna(0).astype(bool)
    )
    df["non_zero_sustainable_investment_share_post_dnsh"] = (
        df["sustainable_investment_share_post_dnsh"].fillna(0).astype(bool)
    )
    for thr in [0.5, 0.33, 0.25, 0.2]:
        df[f"SIS_issuer_binary_{int(thr*100)}"] = df[
            "sustainable_investment_share_post_dnsh"
        ] >= thr
    df["SIS_issuer_binary_05"] = (
        (df["sustainable_investment_share_post_dnsh"] >= 0.05)
        & (df["sustainable_investment_share_post_dnsh"] < 0.2)
    )
    df["SIS_issuer_binary_below_05"] = df["sustainable_investment_share_post_dnsh"] < 0.05
    return df


@register_mastering("dfNZ")
def _nz(df: pd.DataFrame) -> pd.DataFrame:
    alignment_map = {
        "5. Not aligned": "Not Aligned",
        "4. Committed to aligning": "Committed",
        "3. Aligning towards a Net Zero pathway": "Aligning",
        "1. Achieving Net Zero": "Achieving",
    }
    bool_map = {"Validated": True, "Not Validated": False}
    df["nz_alignment_status"] = df["nz_alignment_status"].map(alignment_map)
    for col in ["c1_status", "c2_status", "c4_status", "c5_status"]:
        df[col] = df[col].map(bool_map)
    return df


@register_mastering("dfPAI")
def _pai(df: pd.DataFrame) -> pd.DataFrame:
    df.loc[df["pai_5"] > 1, "pai_5"] = df["pai_5"] / 100
    return df


@register_mastering("dfKPISov")
def _kpisov(df: pd.DataFrame) -> pd.DataFrame:
    # needs lib_utils.sWorkFolder and lib_utils.dfCountry
    trucost = pd.read_excel(Path(sWorkFolder) / "Trucost sovereign dataset.xlsx")
    trucost["Scopes 1 + 2 + 3 (in mn tCO2)"] = trucost[
        ["Scopes 1 + 2 (in mn tCO2)", "Scope 3 (in mn tCO2)"]
    ].sum(axis=1)

    df["carbon_government_ghg_intensity_debt"] = (
        df["carbon_government_ghg"]
        / (df["carbon_government_gdp_nominal_usd"] * df["carbon_government_raw_public_debt"] / 100)
        * 1_000_000
    )
    df = pd.merge(df, dfCountry[["country_code", "country_code3"]], on="country_code", how="left")
    df = pd.merge(df, trucost, left_on="country_code3", right_on="Country ISO", how="left")
    # more KPIs like in original…
    return df


@register_mastering("dfSBTI")
def _sbti(df: pd.DataFrame) -> pd.DataFrame:
    df["near_term_target_year"] = df["near_term_target_year"].apply(extract_max_year)
    df["short_term_targets_2025"] = (
        (df["near_term_target_status"] == "Targets Set") & (df["near_term_target_year"] == 2025)
    )
    df["mid_term_targets_2030-35"] = (
        (df["near_term_target_status"] == "Targets Set")
        & (df["near_term_target_year"].between(2030, 2035))
    )
    df["net_zero_targets"] = df["net_zero_committed"] == "Yes"
    df["carbon_reduction_targets_2030"] = (
        (df["near_term_target_status"] == "Targets Set") & (df["near_term_target_year"] == 2030)
    )
    df["sbti_approved_net_zero_strategy"] = df["long_term_target_status"] == "Targets Set"
    return df


@register_mastering("dfNZAgg")
def _nzagg(df: pd.DataFrame) -> pd.DataFrame:
    nzsov = pd.read_excel(Path(sWorkFolder) / "NZAS sov_2024 11.xlsx", sheet_name="NZAS Sov")
    nzsov = nzsov.drop_duplicates(subset=["id_bb_company"]).dropna(subset=["id_bb_company"])
    nzsov["nz_alignment_status"] = nzsov["nz_alignment_status_sov"].map(
        {
            "Committed to Aligning": "Committed",
            "Not Aligned": "Not Aligned",
            "Aligned to Net Zero": "Aligned",
            "Aligning towards Net Zero": "Aligning",
            "Achieving Net Zero": "Achieving",
        }
    )
    df_out = pd.concat([df, nzsov])
    df_out = df_out.rename(columns={"nz_alignment_status": "nz_alignment_status_agg"})
    return df_out


# -----------------------------------------------------------------------------
# Dispatcher
# -----------------------------------------------------------------------------

def dremio_mastering(dataset: str, df: pd.DataFrame) -> pd.DataFrame:
    """
    Apply dataset-specific mastering rules to a DataFrame.
    """
    if dataset not in MASTERING_TRANSFORMERS:
        raise ValueError(f"No mastering rules registered for dataset {dataset}")
    return MASTERING_TRANSFORMERS[dataset](df)

# dbx
def _infer_date_col(df: pd.DataFrame, fallback: str) -> str:
    """Infer which date column to use."""
    for candidate in ["effective_date", "effective_date_key", "dal_effective_date"]:
        if candidate in df.columns:
            return candidate
    return fallback


def _build_sql(dataset: str, table: str, date_col: str, effective_date: str, hist: bool) -> str:
    """Build Dremio SQL query string."""
    if dataset in {"dfSRI", "dfPSR"} and not hist:
        return f"SELECT * FROM {table} WHERE {date_col} = '{effective_date}'"

    base = f"""
        SELECT f.*, e.id_bb_company as id_bb_company_e, e.party_name
        FROM {table} f
        INNER JOIN {lib_utils.dictDremio['dfEntityDim']} e
          ON e.entity_key = f.entity_key
    """
    condition = f"{date_col} <= '{effective_date}'" if hist else f"{date_col} = '{effective_date}'"
    return f"{base} WHERE {condition}"


def load_dremio_data(
    dataset: str,
    group_by: Optional[Sequence[str]] = None,
    sDate: Optional[str] = None,
    date_col: str = "",
    hist: bool = False,
    mastering: bool = True,
    cache_base_path: str = "dbfs:/Volumes/zenith_dev_gold/esg_impact_analysis/cache_volume/",
) -> pd.DataFrame:
    """
    Load Dremio dataset with caching, mastering, and postprocessing.

    Args:
        dataset: Dataset key (e.g. 'dfSRI').
        group_by: Columns to group by before deduplication.
        sDate: Specific effective date to load (default: latest).
        date_col: Date column to use if not auto-inferred.
        hist: Whether to pull history up to date (True) or just exact match (False).
        mastering: Whether to apply mastering rules.
        cache_base_path: Where cached parquet is stored.

    Returns:
        Mastered and cached pandas DataFrame.
    """
    group_by = list(group_by or [])
    cache_path = f"{cache_base_path}{dataset}.parquet"
    table = lib_utils.dictDremio[dataset]

    # Try cached copy
    try:
        df_local = load_cache(cache_path, lib_utils.spark, format="parquet")
        use_date_col = _infer_date_col(df_local, date_col)
        last_local_date = df_local[use_date_col].astype(str).max()
        logger.info("Cache found", extra={"dataset": dataset, "last_local_date": last_local_date})
    except Exception:
        logger.warning("No cached copy found", extra={"dataset": dataset})
        df_local = None
        use_date_col = date_col
        last_local_date = "1900-01-01"

    # Get latest available date if not specified
    if not sDate:
        if dataset == "dfSusMinExclusions":
            sql_date = f"SELECT max({use_date_col}) as max_date FROM {table} WHERE id_exl_list = 3"
        elif dataset == "dfSusMinExclusionsEnh":
            sql_date = f"SELECT max({use_date_col}) as max_date FROM {table} WHERE id_exl_list = 65"
        else:
            sql_date = f"SELECT max({use_date_col}) as max_date FROM {table}"
        sDate = lib_utils.spark.sql(sql_date).toPandas()["max_date"].astype(str).item()

    # If cache is up-to-date
    if df_local is not None and last_local_date == sDate:
        logger.info("Using cached copy", extra={"dataset": dataset, "date": sDate})
        return df_local

    # Otherwise fetch fresh
    sql = _build_sql(dataset, table, use_date_col, sDate, hist)
    logger.info("Fetching fresh data", extra={"dataset": dataset, "sql": sql})
    df = lib_utils.spark.sql(sql).toPandas()
    df = df.rename(columns=str.lower).apply(pd.to_numeric, errors="ignore")

    # Ensure id_bb_company exists
    if "id_bb_company" not in map(str.lower, df.columns):
        df["id_bb_company"] = df["id_bb_company_e"]

    # Deduplication
    sort_cols = group_by + ["id_bb_company", use_date_col]
    for fallback in ["entity_durable_key", "entity_key", "edm_pty_id"]:
        if fallback in df.columns:
            sort_cols.append(fallback)
            break
    df = df.sort_values(sort_cols).groupby(group_by + ["id_bb_company"]).last().reset_index()

    # Postprocessing
    pct_cols = [c for c in df.columns if c in lib_utils.lVarsDremioPct]
    if pct_cols:
        df = calc_decimal(df, pct_cols)

    if mastering:
        df = dremio_mastering(dataset, df)

    # Save cache
    if df[["id_bb_company"] + group_by].drop_duplicates().shape[0] != len(df):
        raise ValueError("Non-unique id_bb_company after mastering")

    save_cache(df, cache_path, lib_utils.spark, format="parquet")
    return df

def _normalize_weights(df: pd.DataFrame, group_col: str, weight_col: str = "weight") -> pd.DataFrame:
    """Normalize weights so they sum to 1 within each group."""
    df = df.copy()
    df[weight_col] = df[weight_col] / df.groupby(group_col)[weight_col].transform("sum")
    return df


def _map_missing_ids(df: pd.DataFrame, manual_isin: pd.DataFrame, manual_name: pd.DataFrame) -> pd.DataFrame:
    """Fill missing ID_BB_COMPANY using manual mapping tables."""
    df = df.merge(manual_isin, on="id_isin", how="left")
    df["ID_BB_COMPANY"] = df["ID_BB_COMPANY_x"].fillna(df["ID_BB_COMPANY_y"])
    df.drop(columns=["ID_BB_COMPANY_x", "ID_BB_COMPANY_y"], inplace=True)

    df = df.merge(manual_name, on="long_name", how="left")
    df["ID_BB_COMPANY"] = df["ID_BB_COMPANY_x"].fillna(df["ID_BB_COMPANY_y"])
    df.drop(columns=["ID_BB_COMPANY_x", "ID_BB_COMPANY_y"], inplace=True)
    return df


def load_underlyings(
    derivatives_file: str | Path = None,
    manual_mapping_file: str | Path = None,
    effective_date: str = "2024-07-09",
) -> pd.DataFrame:
    """
    Load derivatives underlyings from Excel and merge with securities.

    Args:
        derivatives_file: Path to Derivatives_List.xlsx
        manual_mapping_file: Path to Manual_Mapping.xlsx
        effective_date: Effective date string to stamp into the data.

    Returns:
        DataFrame of underlyings with normalized weights and enriched security info.
    """
    sPathName = lib_utils.sPathNameGeneral
    derivatives_file = Path(derivatives_file or (sPathName + "15. Investment process/Derivatives/Derivatives_List.xlsx"))
    manual_mapping_file = Path(manual_mapping_file or (lib_utils.sWorkFolder + "Manual_Mapping.xlsx"))

    logger.info("Loading derivatives", extra={"file": str(derivatives_file)})

    # base derivatives list
    df_derivatives = pd.read_excel(derivatives_file)

    # external funds
    df_ext = load_holdings("refinitiv", False)
    df_manual_isin = pd.read_excel(manual_mapping_file, sheet_name="id_isin")
    df_manual_name = pd.read_excel(manual_mapping_file, sheet_name="long_name")
    df_ext = _map_missing_ids(df_ext, df_manual_isin, df_manual_name)
    df_ext = df_ext.loc[df_ext.portfolio_id.isin(df_derivatives["ETF ISIN"].unique())]

    # MEMB sheet (equity futures)
    df_bbg = pd.read_excel(derivatives_file, sheet_name="20240709")
    df_bbg["effective_date"] = effective_date
    df_bbg = _normalize_weights(df_bbg, "portfolio_id")
    df_bbg["agi_instrument_asset_type_description"] = "Equity"
    df_bbg = pd.merge(
        df_bbg, lib_utils.dfSecurities, how="left", left_on="id_isin", right_on="ID_ISIN", validate="m:1"
    )
    df_bbg["long_name"] = df_bbg["LONG_COMP_NAME"]
    df_bbg["issuer_long_name"] = df_bbg["LONG_COMP_NAME"]
    df_bbg["portfolio_type"] = "Equity Index"

    # CDX sheet (fixed income futures)
    df_cdx = pd.read_excel(derivatives_file, sheet_name="20240709_CDX")
    df_cdx["effective_date"] = effective_date
    df_cdx = _normalize_weights(df_cdx, "portfolio_id")
    df_cdx["agi_instrument_asset_type_description"] = "Fixed Income"
    df_cdx = pd.merge(
        df_cdx, lib_utils.dfSecurities, how="left", left_on="id_isin", right_on="ID_ISIN", validate="m:1"
    )
    df_cdx["issuer_long_name"] = df_cdx["LONG_COMP_NAME"]
    df_cdx["portfolio_type"] = "CDX"

    # combine
    df_underlyings = pd.concat([df_ext, df_bbg, df_cdx], ignore_index=True)
    df_underlyings["gics_sector_description"] = df_underlyings["INDUSTRY_SECTOR"]

    logger.info("Underlyings loaded", extra={"rows": len(df_underlyings)})
    return df_underlyings

def calc_coverage(
    df_portfolio: pd.DataFrame,
    variables: List[str],
    scope_col: str = "Scope",
) -> pd.DataFrame:
    """
    Calculate coverage flags for a set of variables in a portfolio DataFrame.

    For each variable in `variables`, this adds two columns:
    - Issuer_has_Observation_<var>: True if in scope and variable not null, False if in scope and null, NaN if out of scope.
    - Uncovered_Issuer_<var>: True if in scope but variable missing, False if in scope and available, NaN if out of scope.

    Args:
        df_portfolio: Input portfolio DataFrame.
        variables: List of variable names to check for coverage.
        scope_col: Column indicating whether a row is in scope (default: 'Scope').

    Returns:
        DataFrame with new coverage columns added.
    """
    df_out = df_portfolio.copy()

    for var in variables:
        obs_col = f"Issuer_has_Observation_{var}"
        uncov_col = f"Uncovered_Issuer_{var}"

        # True/False if in scope, NaN if not in scope
        df_out[obs_col] = np.where(
            df_out[scope_col],
            df_out[var].notna(),
            np.nan,
        )

        # True if in scope but missing, False if in scope and present, NaN if out of scope
        df_out[uncov_col] = np.where(
            df_out[scope_col],
            df_out[var].isna(),
            np.nan,
        )

    return df_out

def _apply_lookthrough(
    df: pd.DataFrame,
    lRebased: List[str],
    lRebasedScope: List[str],
    lUnrebased: List[str],
    lUnrebasedScope: List[str],
) -> pd.DataFrame:
    """Handle lookthrough logic for target funds."""
    df = df.copy()
    df["Target Fund Exposure"] = False

    mask = (
        (df.portfolio_type == "Fund")
        & (df.agi_instrument_asset_type_description == "Funds")
        & ~(df.agi_instrument_type_description.fillna("").str.contains("REIT"))
    )
    df.loc[mask, "Target Fund Exposure"] = True
    df.loc[df["Target Fund Exposure"], "Scope"] = True

    return df


def _compute_weighted_sums(
    df: pd.DataFrame,
    metrics: List[str],
    scope_cols: List[str],
    group_by: List[str],
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Compute weighted sums for ESG metrics, masking out-of-scope values."""
    df_out = df.copy()

    # Multiply metrics by weights
    df_out[metrics] = df_out[metrics].multiply(df_out["weight"], axis="index").astype(float)

    # Mask metrics where scope is False
    for metric, scope in zip(metrics, scope_cols):
        df_out.loc[df_out[scope] == False, metric] = np.nan

    # Drop datetime columns (groupby-sum fails with mixed dtypes)
    df_out = df_out.select_dtypes(exclude=["datetime64[ns]"])

    # Aggregate
    df_agg = df_out[group_by + metrics].groupby(group_by).sum()[metrics]
    return df_agg, df_out


def _rebase_metrics(
    df: pd.DataFrame,
    lRebased: List[str],
    lRebasedScope: List[str],
    group_by: List[str],
) -> pd.DataFrame:
    """Rebase rebased metrics by coverage rate."""
    df_rebase = df.copy()

    for metric, scope in zip(lRebased, lRebasedScope):
        df_rebase.loc[df_rebase[scope] == False, metric] = np.nan

    df_rebase = df_rebase.set_index(group_by)
    coverage = (
        df_rebase[lRebased].notnull()
        .multiply(df_rebase["weight"], axis="index")
        .groupby(level=list(range(len(group_by))))
        .sum()
    )
    return coverage


def calc_weighted_average_esg_new(
    dfPortfolio: pd.DataFrame,
    lRebased: List[str],
    lRebasedScope: List[str],
    lUnrebased: List[str],
    lUnrebasedScope: List[str],
    sGroupBy: List[str],
    iLookthrough: bool = False,
    use_equal_weights: bool = False,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Calculate weighted average ESG scores with optional lookthrough.

    Returns:
        dfOutAgg: Aggregated weighted averages per group.
        dfOut: Row-level portfolio with weighted metrics applied.
        dfPortfolio: Original portfolio (possibly modified with lookthrough logic).
    """
    if iLookthrough:
        dfPortfolio = _apply_lookthrough(dfPortfolio, lRebased, lRebasedScope, lUnrebased, lUnrebasedScope)

    # All metrics
    metrics = lRebased + lUnrebased
    scope_cols = lRebasedScope + lUnrebasedScope

    # Weighted sums
    dfOutAgg, dfOut = _compute_weighted_sums(dfPortfolio, metrics, scope_cols, sGroupBy)

    # Rebasing
    coverage = _rebase_metrics(dfPortfolio, lRebased, lRebasedScope, sGroupBy)
    dfOutAgg[lRebased] = dfOutAgg[lRebased] / coverage[lRebased]

    # Equal-weight alternative
    if use_equal_weights:
        raise NotImplementedError("Equal-weight logic not yet migrated from legacy block.")

    return dfOutAgg, dfOut, dfPortfolio
   

def _prepare_esg_data(dfESG: pd.DataFrame, sRightOn: str, sVariables: Optional[List[str]]) -> pd.DataFrame:
    """Prepare ESG dataset: rename, filter variables, drop NAs, validate uniqueness."""
    dfESG = dfESG.copy(deep=True)
    dfESG.rename(columns={"id_bb_company": "ID_BB_COMPANY"}, inplace=True)

    if sVariables:
        dfESG = dfESG[[sRightOn] + sVariables]
        dfESG.dropna(subset=[sVariables[0]], inplace=True)

    if dfESG[sRightOn].nunique() != len(dfESG):
        raise ValueError("ESG dataset is not unique by right key")

    return dfESG


def _merge_esg_core(
    dfPortfolio: pd.DataFrame,
    dfESG: pd.DataFrame,
    sVariables: Optional[List[str]],
    sESGSuffix: str,
    iWaterfall: bool,
    sLeftOn: str,
    sRightOn: str,
    sScope: str,
) -> pd.DataFrame:
    """Perform the core merge depending on waterfall or direct join."""
    if not iWaterfall:
        if sVariables:
            dfOut = pd.merge(
                dfPortfolio,
                dfESG.assign(Scope=True),
                how="left",
                left_on=[sScope, sLeftOn],
                right_on=["Scope", sRightOn],
                suffixes=("", sESGSuffix),
                validate="m:1",
            )
        else:
            dfOut = pd.merge(
                dfPortfolio,
                dfESG.assign(Scope=True),
                how="left",
                left_on=[sScope, sLeftOn],
                right_on=["Scope", sRightOn],
                suffixes=("", sESGSuffix),
                indicator="Exposure" + sESGSuffix,
                validate="m:1",
            )
            dfOut["Exposure" + sESGSuffix] = dfOut["Exposure" + sESGSuffix].map({"both": True, "left_only": False})
            dfOut.loc[dfOut[sScope] == False, "Exposure" + sESGSuffix] = np.nan
    else:
        # waterfall case
        dfOut = calc_waterfall(dfPortfolio, dfESG, sESGSuffix + "_used")
        merge_key = sLeftOn + sESGSuffix + "_used"

        dfOut = pd.merge(
            dfOut,
            dfESG.assign(Scope=True),
            how="left",
            left_on=[sScope, merge_key],
            right_on=["Scope", sRightOn],
            suffixes=("", sESGSuffix),
            validate="m:1",
        )
    return dfOut


def _apply_suffix_logic(dfOut: pd.DataFrame, sESGSuffix: str, sVariables: Optional[List[str]]) -> pd.DataFrame:
    """Post-processing depending on ESG suffix or variable."""
    if sESGSuffix == "_FH":
        dfOut["FH"] = (dfOut["Sovereigns"] == 1) & (dfOut["status"] == 1)
        dfOut["status"] = np.where(dfOut["FH"], True, np.where(dfOut["Scope"] == False, np.nan, False))

    elif sESGSuffix == "_GB":
        mask = (dfOut["Exposure_GB"] == 1) & (dfOut["dnsh_flag_overall"] == False)
        dfOut.loc[mask, ["sustainable_investment_share_pre_dnsh", "sustainable_investment_share_post_dnsh"]] = 1

    elif sVariables == ["GREEN_BOND_LOAN_INDICATOR"]:
        mask = (dfOut["GREEN_BOND_LOAN_INDICATOR"] == 1) & (dfOut["dnsh_flag_overall"] == False)
        dfOut.loc[mask, ["sustainable_investment_share_environmental_post_dnsh"]] = 1
        dfOut.loc[mask, ["sustainable_investment_share_social_post_dnsh"]] = 0

    elif sVariables == ["SOCIAL_BOND_IND"]:
        mask = (dfOut["SOCIAL_BOND_IND"] == 1) & (dfOut["dnsh_flag_overall"] == False)
        dfOut.loc[mask, ["sustainable_investment_share_social_post_dnsh"]] = 1
        dfOut.loc[mask, ["sustainable_investment_share_environmental_post_dnsh"]] = 0

    elif sVariables == ["SUSTAINABILITY_BOND_IND"]:
        mask = (dfOut["SUSTAINABILITY_BOND_IND"] == 1) & (dfOut["dnsh_flag_overall"] == False)
        dfOut.loc[mask, ["sustainable_investment_share_social_post_dnsh"]] = 0.5
        dfOut.loc[mask, ["sustainable_investment_share_environmental_post_dnsh"]] = 0.5

    elif sVariables == ["CO2_Engaged"]:
        dfOut["CO2_Engaged"] = 0
        dfOut.loc[dfOut["Exposure_Eng"] == True, "CO2_Engaged"] = dfOut["carbon_emissions_scope_12"]

    elif sESGSuffix == "_Ex":
        dfOut[dfOut.columns.difference(["ID_BB_COMPANY"])] = dfOut.drop(columns=["ID_BB_COMPANY"]).fillna(False)

    elif sESGSuffix == "_Comp":
        dfOut["Exposure_Comp"] = dfOut[["Exposure_BS", "status"]].max(axis=1)

    return dfOut


def merge_esg_data_new(
    dfPortfolios: pd.DataFrame,
    dfESGIn: pd.DataFrame,
    sVariables: Optional[List[str]],
    sESGSuffix: str,
    iWaterfall: bool = False,
    sLeftOn: str = "ID_BB_COMPANY",
    sRightOn: str = "ID_BB_COMPANY",
    sScope: str = "Scope",
) -> pd.DataFrame:
    """
    Merge portfolio data with ESG dataset, applying suffix-specific post-processing.
    """
    dfPortfolio = dfPortfolios.copy()
    dfESG = _prepare_esg_data(dfESGIn, sRightOn, sVariables)
    dfOut = _merge_esg_core(dfPortfolio, dfESG, sVariables, sESGSuffix, iWaterfall, sLeftOn, sRightOn, sScope)
    dfOut = _apply_suffix_logic(dfOut, sESGSuffix, sVariables)
    return dfOut


def calc_new_variables(dfPortfolios):
    dfPortfolios['Uncovered_Issuer_AGI Custom ESMA PAB Results - 2024 November_exSRI'] = False
    
    dfPortfolios['AGI Custom ESMA PAB Results - 2024 November_exSRI'] = False
    dfPortfolios.loc[(dfPortfolios.Exposure_Sus == False) & (dfPortfolios['AGI Custom ESMA PAB Results - 2024 November'] == True),'AGI Custom ESMA PAB Results - 2024 November_exSRI'] = True
    

    # dfPortfolios['Worst_Quintile_PSR'] = (dfPortfolios['PSR_score_final_relative_postFlags'] 
    #                                   <= dfPortfolios.drop_duplicates(subset=['ID_BB_COMPANY'])['PSR_score_final_relative_postFlags'].quantile(0.2)
    #                                   ).where(dfPortfolios['PSR_score_final_relative_postFlags'].notna())
    # dfPortfolios['Worst_Quintile_SRI'] = (dfPortfolios['SRI_Rating_Final'] 
    #                                   <= dfPortfolios.drop_duplicates(subset=['ID_BB_COMPANY'])['SRI_Rating_Final'].quantile(0.2)
    #                                   ).where(dfPortfolios['SRI_Rating_Final'].notna())

    # #dfPortfolios['Worst_Quintile_SRI'] = (dfPortfolios['SRI_Rating_Final']<= dfPortfolios.drop_duplicates(subset = ['ID_BB_COMPANY'])['SRI_Rating_Final'].quantile(0.2)).astype(bool)
    # dfPortfolios['Downgrades'] = ((dfPortfolios['Worst_Quintile_SRI'] == False) & 
    #                           (dfPortfolios['Worst_Quintile_PSR'] == True)).astype(bool)
    # dfPortfolios['Upgrades'] = ((dfPortfolios['Worst_Quintile_SRI'] == True) & 
    #                           (dfPortfolios['Worst_Quintile_PSR'] == False)).astype(bool)

    # dfPortfolios['Align_Add_Cov_MSCI'] = dfPortfolios['Uncovered_Issuer_nz_alignment_status'].astype(bool) & ~dfPortfolios['Uncovered_Issuer_PAII_NZIF_ALIGNMENT'].astype(bool)
    # dfPortfolios['Align_Lost_Cov_MSCI'] = ~dfPortfolios['Uncovered_Issuer_nz_alignment_status'].astype(bool) & dfPortfolios['Uncovered_Issuer_PAII_NZIF_ALIGNMENT'].astype(bool)
    # dfPortfolios['Align_Overlap_Cov_MSCI'] = ~dfPortfolios['Uncovered_Issuer_nz_alignment_status'].astype(bool) & ~dfPortfolios['Uncovered_Issuer_PAII_NZIF_ALIGNMENT'].astype(bool)

    # dfPortfolios['c1_Add_Cov_MSCI'] = dfPortfolios['Uncovered_Issuer_c1_status'].astype(bool) & ~dfPortfolios['Uncovered_Issuer_PAII_NZIF_CRITERIA_1'].astype(bool)
    # dfPortfolios['c1_Lost_Cov_MSCI'] = ~dfPortfolios['Uncovered_Issuer_c1_status'].astype(bool) & dfPortfolios['Uncovered_Issuer_PAII_NZIF_CRITERIA_1'].astype(bool)
    # dfPortfolios['c1_Overlap_Cov_MSCI'] = ~dfPortfolios['Uncovered_Issuer_c1_status'].astype(bool) & ~dfPortfolios['Uncovered_Issuer_PAII_NZIF_CRITERIA_1'].astype(bool)

    # dfPortfolios['c2_Add_Cov_MSCI'] = dfPortfolios['Uncovered_Issuer_c2_status'].astype(bool) & ~dfPortfolios['Uncovered_Issuer_PAII_NZIF_CRITERIA_2'].astype(bool)
    # dfPortfolios['c2_Lost_Cov_MSCI'] = ~dfPortfolios['Uncovered_Issuer_c2_status'].astype(bool) & dfPortfolios['Uncovered_Issuer_PAII_NZIF_CRITERIA_2'].astype(bool)
    # dfPortfolios['c2_Overlap_Cov_MSCI'] = ~dfPortfolios['Uncovered_Issuer_c2_status'].astype(bool) & ~dfPortfolios['Uncovered_Issuer_PAII_NZIF_CRITERIA_2'].astype(bool)
    
    # dfPortfolios['c4_Add_Cov_MSCI'] = dfPortfolios['Uncovered_Issuer_c4_status'].astype(bool) & ~dfPortfolios['Uncovered_Issuer_PAII_NZIF_CRITERIA_4'].astype(bool)
    # dfPortfolios['c4_Lost_Cov_MSCI'] = ~dfPortfolios['Uncovered_Issuer_c4_status'].astype(bool) & dfPortfolios['Uncovered_Issuer_PAII_NZIF_CRITERIA_4'].astype(bool)
    # dfPortfolios['c4_Overlap_Cov_MSCI'] = ~dfPortfolios['Uncovered_Issuer_c4_status'].astype(bool) & ~dfPortfolios['Uncovered_Issuer_PAII_NZIF_CRITERIA_4'].astype(bool)

    # dfPortfolios['c5_Add_Cov_MSCI'] = dfPortfolios['Uncovered_Issuer_c5_status'].astype(bool) & ~dfPortfolios['Uncovered_Issuer_PAII_NZIF_CRITERIA_5'].astype(bool)
    # dfPortfolios['c5_Lost_Cov_MSCI'] = ~dfPortfolios['Uncovered_Issuer_c5_status'].astype(bool) & dfPortfolios['Uncovered_Issuer_PAII_NZIF_CRITERIA_5'].astype(bool)
    # dfPortfolios['c5_Overlap_Cov_MSCI'] = ~dfPortfolios['Uncovered_Issuer_c5_status'].astype(bool) & ~dfPortfolios['Uncovered_Issuer_PAII_NZIF_CRITERIA_5'].astype(bool)
    
    # dfPortfolios['Align_match'] = dfPortfolios['nz_alignment_status'] == dfPortfolios['PAII_NZIF_ALIGNMENT']
    # dfPortfolios['c1_match'] = dfPortfolios['c1_status'] == dfPortfolios['PAII_NZIF_CRITERIA_1']
    # dfPortfolios['c2_match'] = dfPortfolios['c2_status'] == dfPortfolios['PAII_NZIF_CRITERIA_2']
    # dfPortfolios['c4_match'] = dfPortfolios['c4_status'] == dfPortfolios['PAII_NZIF_CRITERIA_4']
    # dfPortfolios['c5_match'] = dfPortfolios['c5_status'] == dfPortfolios['PAII_NZIF_CRITERIA_5']

    # dfPortfolios['carbon_emissions_scope_123_total_sales_inten'] = dfPortfolios['carbon_emissions_scope_3_total_sales_inten'] + dfPortfolios['carbon_emissions_scope_12_inten']
    # dfPortfolios['carbon_emissions_scope_123_total_evic_inten'] = dfPortfolios['carbon_emissions_scope_3_tot_evic_inten'] + dfPortfolios['carbon_emissions_evic_scope_12_inten']
    #dfPortfolios['apportioned_revenues'] = dfPortfolios['sales_usd_recent']*0.9/dfPortfolios['evic_eur']
    #dfPortfolios['ethica'] = dfPortfolios[['AGI Custom Ethica_2025 March','sri_rating_final < 2','Exposure_Sus']].any(axis=1)
    # print('test')
    # quantiles = dfPortfolios.groupby(['gics_sector_description','country_name'])['industry_adjusted_score'].quantile(0.3,interpolation = 'lower').reset_index()
    # quantiles.columns = ['gics_sector_description'  ,  'country_name',  'industry_adjusted_score_']
    # dfPortfolios = pd.merge(dfPortfolios,quantiles,how = 'left',on = ['gics_sector_description','country_name'])
    # dfPortfolios['is_worst_30'] = False
    # dfPortfolios.loc[dfPortfolios['industry_adjusted_score']<dfPortfolios['industry_adjusted_score_'],'is_worst_30'] = True
    # print(dfPortfolios['is_worst_30'])
    #dfPortfolio['carbon_emissions_scope_12_inten_apportioned_sales'] = dfPortfolios['']
    # dfPortfolios['PAB_sustainable_investment_share_post_dnsh'] = dfPortfolios['sustainable_investment_share_post_dnsh']
    #dfPortfolios.loc[dfPortfolios['AGI Custom ESMA PAB (CW RED)_Issuers_20240117'] == True,'PAB_sustainable_investment_share_post_dnsh'] = 0
    
    # dfPortfolios['region_DM'] = False
    # dfPortfolios.loc[dfPortfolios.country_name.isin(['Germany',
    #  'Denmark',
    #  'Sweden',
    #  'Switzerland',
    #  'Netherlands',
    #  'Hong Kong',
    #  'United Kingdom',
    #  'Singapore',
    #  'Norway',
    #  'United States',
    #  'Australia',
    #  'France',
    #  'Japan',
    #  'New Zealand',
    #  'Ireland',
    #  'Luxembourg',
    #  'Spain',
    #  'Canada',
    #  'Italy',
    #  'Belgium',
    #  'Israel',
    #  'Finland',
    #  'Austria',
    #  'Portugal']),'region_DM'] = True
    # dfPortfolios['region_EM'] = ~dfPortfolios['region_DM']
    # dfPortfolios['Green instrument'] = ''
    # dfPortfolios.loc[(dfPortfolios.Scope_Inv == True)&(dfPortfolios['region'] == 'DM')&(dfPortfolios['esg_rating'].isin(['AAA','AA','A','BBB'])),'Green instrument'] = True
    # dfPortfolios.loc[(dfPortfolios.Scope_Inv == True)&(dfPortfolios['region'] == 'EM')&(dfPortfolios['esg_rating'].isin(['AAA','AA','A','BBB','BB'])),'Green instrument'] = True
    #dfPortfolios['SI_NZ_share'] = dfPortfolios['sustainable_investment_share_post_dnsh']
    #dfPortfolios.loc[(dfPortfolios['dnsh_flag_overall']== False)&
    #                (dfPortfolios.nz_alignment_status.isin(['3. Aligning towards a Net Zero pathway','1. Achieving Net Zero'])),'SI_NZ_share'] = 1
    #dfPortfolios['total_exclusions'] = dfPortfolios[['Exposure_Sus','Exposure_TS','Exposure_Febelfin','Exposure_Febelfin25','Exposure_Urgewald','Exposure_GOGEL','Exposure_GCEL']].max(axis = 1)
    # dfPortfolios['Target Fund Exposure'] = 0
    # dfPortfolios['Score downgrade < 2'] = False
    # dfPortfolios['Score upgrade > 2'] = False
    # dfPortfolios['Score downgrade < 1'] = False
    # dfPortfolios['Score upgrade > 1'] = False
    # # dfPortfolios['Rating downgrade < E'] = False
    # # dfPortfolios['Rating upgrade > E'] = False
    # # dfPortfolios['Rating downgrade < D'] = False
    # # dfPortfolios['Rating upgrade > D'] = False
    # dfPortfolios.loc[(dfPortfolios['SRI_Rating_Final']>=2)&(dfPortfolios['PSR_score_final_relative_postFlags']<2),'Score downgrade < 2'] = True
    # dfPortfolios.loc[(dfPortfolios['SRI_Rating_Final']<2)&(dfPortfolios['PSR_score_final_relative_postFlags']>=2),'Score upgrade > 2'] = True
    # dfPortfolios.loc[(dfPortfolios['SRI_Rating_Final']>=1)&(dfPortfolios['PSR_score_final_relative_postFlags']<1),'Score downgrade < 1'] = True
    # dfPortfolios.loc[(dfPortfolios['SRI_Rating_Final']<1)&(dfPortfolios['PSR_score_final_relative_postFlags']>=1),'Score upgrade > 1'] = True
    # dfPortfolios.loc[(dfPortfolios['sri_rating_final']>=1)&(dfPortfolios['PSR_rating_final'].isin(['E'])),'Rating downgrade < E'] = True
    # dfPortfolios.loc[(dfPortfolios['sri_rating_final']<1)&(dfPortfolios['PSR_rating_final'].isin(['A','B','C','D'])),'Rating upgrade > E'] = True
    # dfPortfolios.loc[(dfPortfolios['sri_rating_final']>=2)&(dfPortfolios['PSR_rating_final'].isin(['D','E'])),'Rating downgrade < D'] = True
    # dfPortfolios.loc[(dfPortfolios['sri_rating_final']<2)&(dfPortfolios['PSR_rating_final'].isin(['A','B','C'])),'Rating upgrade > D'] = True
    # dfPortfolios['Coverage_Lost'] = False
    # dfPortfolios.loc[(dfPortfolios.Sovereigns == False)&(dfPortfolios.PSR_score_final_raw.isna())&(~dfPortfolios.sri_rating_final.isna()),'Coverage_Lost'] = True
    # dfPortfolios['Coverage_Gained'] = False
    # dfPortfolios.loc[(dfPortfolios.Sovereigns == False)&(~dfPortfolios.PSR_score_final_raw.isna())&(dfPortfolios.sri_rating_final.isna()),'Coverage_Gained'] = True
    # dfPortfolios['Scope'] = dfPortfolios['Scope_Corp']
    return dfPortfolios

def calc_waterfall(
    dfPortfolio: pd.DataFrame,
    dfESG: pd.DataFrame,
    sESGSuffix: str,
    hierarchy_cols: list[str] = None
) -> pd.DataFrame:
    """
    Apply a waterfall match: start with company, fall back to parent, then ultimate parent.

    Args:
        dfPortfolio: Portfolio dataframe with issuer hierarchy columns.
        dfESG: ESG dataframe with available IDs.
        sESGSuffix: Suffix to append to generated ID columns.
        hierarchy_cols: Ordered list of columns to use in the fallback hierarchy.
                        Defaults to company → parent → ultimate parent.

    Returns:
        pd.DataFrame: Portfolio dataframe with resolved ESG IDs.
    """
    if hierarchy_cols is None:
        hierarchy_cols = [
            "ID_BB_COMPANY",
            "ID_BB_PARENT_CO",
            "ID_BB_ULTIMATE_PARENT_CO",
        ]

    dfOut = dfPortfolio.copy()
    valid_ids = set(dfESG["ID_BB_COMPANY"].dropna().unique())

    temp_cols = []
    for col in hierarchy_cols:
        temp_col = f"{col}{sESGSuffix}"
        dfOut[temp_col] = dfOut[col].mask(~dfOut[col].isin(valid_ids))
        temp_cols.append(temp_col)

    # Final waterfall result: take first non-null across hierarchy
    dfOut[f"ID_BB_COMPANY{sESGSuffix}"] = dfOut[temp_cols].bfill(axis=1).iloc[:, 0]

    # Drop intermediate temp cols
    dfOut = dfOut.drop(columns=temp_cols[1:])  # keep the final "ID_BB_COMPANY_suffix"

    return dfOut
# load holdings
def load_portfolio_benchmark_bridge(
    lPortfolioID: List[int],
    con=lib_utils.conndremio,
    dictDremio: dict = lib_utils.dictDremio,
) -> pd.DataFrame:
    """
    Load benchmark bridge data for given portfolio IDs.

    Args:
        lPortfolioID: List of portfolio IDs to filter.
        con: Database connection (default: lib_utils.conndremio).
        dictDremio: Dictionary mapping dataset names to table paths.

    Returns:
        DataFrame with portfolio-to-benchmark mappings.
    """
    if not lPortfolioID:
        logger.warning("No portfolio IDs provided. Returning empty DataFrame.")
        return pd.DataFrame()

    # Build parameterized query for safety
    placeholders = ",".join(["%s"] * len(lPortfolioID))
    sql = f"""
        SELECT b.portfolio_id,
               b.portfolio_name AS portfolio_benchmark_group_name,
               a.benchmark_durable_key,
               c.benchmark_name
        FROM   {dictDremio['dfPortfolioBenchmarkBridge']} a
        LEFT JOIN {dictDremio['dfPortfolioDim']} b
               ON a.portfolio_durable_key = b.portfolio_durable_key
        LEFT JOIN {dictDremio['dfBenchmarkDim']} c
               ON a.benchmark_durable_key = c.benchmark_durable_key
        WHERE  b.is_current = TRUE
          AND  c.is_current = TRUE
          AND  b.portfolio_id IN ({placeholders})
    """

    logger.debug("Executing SQL for portfolio benchmark bridge", extra={"sql": sql, "ids": lPortfolioID})

    df = pd.read_sql(sql, con, params=lPortfolioID)

    return df

# -----------------------------
# Small helpers
# -----------------------------

def _build_in_clause(values: Union[Iterable, int, str]) -> str:
    """
    Build a SQL IN (...) clause list for ints/strings/dates.
    Example: ["2024-01-01", "2024-01-02"] -> "('2024-01-01','2024-01-02')"
             [1, 2] -> "(1, 2)"
    """
    if not isinstance(values, (list, tuple, set)):
        values = [values]

    vals: List[str] = []
    for v in values:
        if v is None:
            continue
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            vals.append(str(int(v)))
        else:
            s = str(v).replace("'", "''")  # rudimentary escaping for quotes
            vals.append(f"'{s}'")
    if not vals:
        # Avoid invalid SQL like IN ()
        vals = ["NULL"]
    return f"({', '.join(vals)})"


# You provided this in another snippet; assumed available:
# def load_portfolio_benchmark_bridge(lPortfolioID): ...

# You also referenced these objects from lib_utils:
# - lib_utils.conndremio (DB connection)
# - lib_utils.dictDremio (table map)
# - lib_utils.dfHierarchy (BBG company hierarchy)
# - lib_utils.dfCountry, lib_utils.dfSecurities (for XLS enrichment)
# - lib_utils.EQ_FI_SPLIT_IDS (config for sleeve splitting)
# - split_eq_fi_sleeves(...) (function), assumed available


# -----------------------------
# Holdings (portfolios)
# -----------------------------

def _load_dremio_holdings(
    lPortfolioIDs: Sequence[Union[int, str]],
    sDate: Sequence[str],
    position_type: int = 4,
    iBenchmarks: bool = False,
    split_ids: Optional[dict] = None,
) -> Union[pd.DataFrame, Tuple[pd.DataFrame, pd.DataFrame]]:
    """
    Load portfolio holdings from Dremio (DAL).

    Args:
        lPortfolioIDs: List of CEDAR portfolio_ids.
        sDate: One or more 'YYYY-MM-DD' dates.
        position_type: GIDP position type key (1..5). Default 4 = Traded/EOD.
        iBenchmarks: If True, also return the benchmark bridge as second DataFrame.
        split_ids: Optional split config passed to split_eq_fi_sleeves.

    Returns:
        dfHoldings or (dfHoldings, dfBenchmarkBridge) if iBenchmarks=True
    """
    ids_clause = _build_in_clause(lPortfolioIDs)
    dates_clause = _build_in_clause(sDate)

    query = f"""
    SELECT
        h.effective_date,
        'Fund' AS portfolio_type,
        p.portfolio_durable_key,
        p.portfolio_id,
        p.portfolio_name,
        p.gidp_reporting_location,
        -- p.prod_line_name_long,
        p.investment_region,
        p.sri_esg,
        p.asset_class,
        p.investment_approach,
        p.market_cap,
        n.market_value_clean_portfolio,
        n.market_value_clean_firm,
        n.portfolio_currency_code,
        v.eu_sfdr_sustainability_category,
        v.internal_asset_class,
        v.domicile,
        h.holding_type,
        h.percentage_of_market_value_clean_portfolio / 100 AS weight,
        h.futures_exposure_firm / n.market_value_clean_firm AS futures_exposure_firm,
        s.country_issue_name,
        e.country_name,
        e.country_code3,
        s.currency_traded,
        'NA' AS ml_high_yield_sector,
        s.agi_instrument_asset_type_description,
        s.security_durable_key,
        s.id_underlying_durable_key,
        s.issuer_long_name,
        s.long_name,
        s.agi_instrument_type_description,
        g.GICS AS gics_sector_description,
        bics.BICS AS bics_sector_description,
        b3.BICS AS bics_l3_description,
        b4.BBG AS BBG_l2_description,
        s.collat_type,
        'NA' AS gics_industry_group_description,
        'NA' AS gics_industry_description,
        s.bloomberg_security_typ,
        s.id_isin,
        s.id_cusip,
        s.security_des,
        s.issue_description,
        s.issuer_id_bloomberg_company AS ID_BB_COMPANY
    FROM {lib_utils.dictDremio['dfHoldings']} h
    JOIN {lib_utils.dictDremio['dfPortfolioDim']} p
      ON p.portfolio_durable_key = h.portfolio_durable_key
    LEFT JOIN {lib_utils.dictDremio['dfSecurityMasterDim']} s
      ON s.security_durable_key = h.security_durable_key
     AND s.is_current = TRUE
    LEFT JOIN {lib_utils.dictDremio['dfCountryDim']} e
      ON s.country_issue = e.country_code
    LEFT JOIN {lib_utils.dictDremio['dfPortfolioTotal']} n
      ON p.portfolio_durable_key = n.portfolio_durable_key
     AND n.date_key = h.date_key
     AND n.position_type_key = h.position_type_key
    LEFT JOIN {lib_utils.dictDremio['dfVehicleDim']} v
      ON p.portfolio_durable_key = v.portfolio_durable_key
     AND v.is_current = TRUE
     AND v.vehicle_type_level = 'vehicle'
    LEFT JOIN (
        SELECT DISTINCT a.security_durable_key, l1_desc AS GICS
        FROM {lib_utils.dictDremio['dfSecuritySector']} a
        LEFT JOIN {lib_utils.dictDremio['dfSecurityMaster']} b
          ON a.security_durable_key = b.security_durable_key
        WHERE a.sector_scheme = 'IDS_GICS'
          AND a.is_current = TRUE
          AND a.security_durable_key IS NOT NULL
    ) g ON s.security_durable_key = g.security_durable_key
    LEFT JOIN (
        SELECT DISTINCT a.security_durable_key, l1_desc AS BICS
        FROM {lib_utils.dictDremio['dfSecuritySector']} a
        LEFT JOIN {lib_utils.dictDremio['dfSecurityMaster']} b
          ON a.security_durable_key = b.security_durable_key
        WHERE a.sector_scheme = 'IDS_BICS'
          AND a.is_current = TRUE
          AND a.security_durable_key IS NOT NULL
    ) bics ON s.security_durable_key = bics.security_durable_key
    LEFT JOIN (
        SELECT DISTINCT a.security_durable_key, l3_desc AS BICS
        FROM {lib_utils.dictDremio['dfSecuritySector']} a
        LEFT JOIN {lib_utils.dictDremio['dfSecurityMaster']} b
          ON a.security_durable_key = b.security_durable_key
        WHERE a.sector_scheme = 'IDS_BICS'
          AND a.is_current = TRUE
          AND a.security_durable_key IS NOT NULL
    ) b3 ON s.security_durable_key = b3.security_durable_key
    LEFT JOIN (
        SELECT DISTINCT a.security_durable_key, l2_desc AS BBG
        FROM {lib_utils.dictDremio['dfSecuritySector']} a
        LEFT JOIN {lib_utils.dictDremio['dfSecurityMaster']} b
          ON a.security_durable_key = b.security_durable_key
        WHERE a.sector_scheme = 'IDS_BBG'
          AND a.is_current = TRUE
          AND a.security_durable_key IS NOT NULL
    ) b4 ON s.security_durable_key = b4.security_durable_key
    WHERE h.effective_date IN {dates_clause}
      AND p.portfolio_id IN {ids_clause}
      AND h.position_type_key = {position_type}
      AND p.is_current = TRUE
      AND p.cluster_type <> 'Operations-Only Portfolio'
      AND p.portfolio_status <> 'Terminated'
      AND p.hierarchy_type NOT IN ('Void')
    """

    print('*********************************************')
    print('Loading holdings data as of', dates_clause, '...')
    print('*********************************************')
    print(query)

    dfHoldings = pd.read_sql(query, lib_utils.conndremio)

    # Attach benchmark bridge and (optionally) fetch benchmark legs
    dfBenchmarkBridge = load_portfolio_benchmark_bridge(lPortfolioIDs)

    if len(dfBenchmarkBridge) > 0:
        dfHoldings = pd.merge(
            dfHoldings,
            dfBenchmarkBridge,
            how="left",
            on="portfolio_id",
            validate="m:1",
        )
        dfHoldings["benchmark_id"] = dfHoldings["benchmark_durable_key"]
        dfHoldings = dfHoldings.drop(columns=["benchmark_durable_key"], errors="ignore")

        lBenchmarkIDs = dfBenchmarkBridge.benchmark_durable_key.drop_duplicates().to_list()
        if len(lBenchmarkIDs) == 1:
            lBenchmarkIDs = [lBenchmarkIDs[0]]

        # Fetch benchmark holdings and append
        dfAllBenchmarks = _load_benchmark_holdings(lBenchmarkIDs, sDate)
        dfAllBenchmarks["portfolio_id"] = dfAllBenchmarks["benchmark_id"]
        dfHoldings = pd.concat([dfHoldings, dfAllBenchmarks], ignore_index=True)

    # Enrich with hierarchy
    dfHoldings = pd.merge(
        dfHoldings,
        lib_utils.dfHierarchy,
        how="left",
        on="ID_BB_COMPANY",
        validate="m:1",
    )

    # Optional sleeve split
    if split_ids is None:
        split_ids = lib_utils.EQ_FI_SPLIT_IDS
    if split_ids:
        dfHoldings = split_eq_fi_sleeves(dfHoldings, split_ids)

    if iBenchmarks:
        return dfHoldings, dfBenchmarkBridge
    return dfHoldings


# -----------------------------
# Benchmarks
# -----------------------------

def _load_benchmark_holdings(
    ids: Union[int, Sequence[int]],
    sDate: Sequence[str],
    position_type: str = "C",
) -> pd.DataFrame:
    """
    Load benchmark holdings from Dremio.

    Args:
        ids: One or more benchmark_durable_key values.
        sDate: One or more 'YYYY-MM-DD' dates.
        position_type: 'C' (close) or 'O' (open). (The current DAL field used is weight_close.)

    Returns:
        DataFrame of benchmark holdings.
    """
    ids_clause = _build_in_clause(ids)
    dates_clause = _build_in_clause(sDate)

    query = f"""
    SELECT
        CAST(d.full_date AS DATE) AS effective_date,
        'Benchmark' AS portfolio_type,
        'benchmark_durable_key' AS durable_key_type,
        b.benchmark_durable_key AS durable_key,
        'benchmark_key' AS key_type,
        b.benchmark_key AS key,
        b.benchmark_durable_key AS benchmark_id,
        b.benchmark_name AS portfolio_name,
        b.benchmark_name AS benchmark_name,
        b.class AS asset_class,
        s.country_issue_name,
        e.country_name,
        e.country_code3,
        s.currency_traded,
        'NA' AS ml_high_yield_sector,
        s.agi_instrument_asset_type_description,
        s.security_durable_key,
        s.long_name,
        s.issuer_long_name,
        s.agi_instrument_type_description,
        g.GICS AS gics_sector_description,
        bics.BICS AS bics_sector_description,
        b3.BICS AS bics_l3_description,
        b4.BBG AS BBG_l2_description,
        s.collat_type,
        'NA' AS gics_industry_group_description,
        'NA' AS gics_industry_description,
        s.bloomberg_security_typ,
        s.id_isin,
        s.id_cusip,
        s.security_des,
        s.issue_description,
        s.issuer_id_bloomberg_company AS ID_BB_COMPANY,
        h.weight_close / 100 AS weight
    FROM {lib_utils.dictDremio['dfBenchmarkConst']} h
    JOIN {lib_utils.dictDremio['dfBenchmarkDim']} b
      ON b.benchmark_durable_key = h.benchmark_durable_key
     AND b.is_current = TRUE
    JOIN {lib_utils.dictDremio['dfDateDim']} d
      ON h.date_key = d.date_key
    LEFT JOIN {lib_utils.dictDremio['dfSecurityMasterDim']} s
      ON s.security_durable_key = h.security_durable_key
     AND s.is_current = TRUE
    LEFT JOIN {lib_utils.dictDremio['dfCountryDim']} e
      ON e.country_code = s.country_issue
    LEFT JOIN (
        SELECT DISTINCT a.security_durable_key, l1_desc AS GICS
        FROM {lib_utils.dictDremio['dfSecuritySector']} a
        LEFT JOIN {lib_utils.dictDremio['dfSecurityMaster']} b
          ON a.security_durable_key = b.security_durable_key
        WHERE a.sector_scheme = 'IDS_GICS'
          AND a.is_current = TRUE
          AND a.security_durable_key IS NOT NULL
    ) g ON s.security_durable_key = g.security_durable_key
    LEFT JOIN (
        SELECT DISTINCT a.security_durable_key, l1_desc AS BICS
        FROM {lib_utils.dictDremio['dfSecuritySector']} a
        LEFT JOIN {lib_utils.dictDremio['dfSecurityMaster']} b
          ON a.security_durable_key = b.security_durable_key
        WHERE a.sector_scheme = 'IDS_BICS'
          AND a.is_current = TRUE
          AND a.security_durable_key IS NOT NULL
    ) bics ON s.security_durable_key = bics.security_durable_key
    LEFT JOIN (
        SELECT DISTINCT a.security_durable_key, l3_desc AS BICS
        FROM {lib_utils.dictDremio['dfSecuritySector']} a
        LEFT JOIN {lib_utils.dictDremio['dfSecurityMaster']} b
          ON a.security_durable_key = b.security_durable_key
        WHERE a.sector_scheme = 'IDS_BICS'
          AND a.is_current = TRUE
          AND a.security_durable_key IS NOT NULL
    ) b3 ON s.security_durable_key = b3.security_durable_key
    LEFT JOIN (
        SELECT DISTINCT a.security_durable_key, l2_desc AS BBG
        FROM {lib_utils.dictDremio['dfSecuritySector']} a
        LEFT JOIN {lib_utils.dictDremio['dfSecurityMaster']} b
          ON a.security_durable_key = b.security_durable_key
        WHERE a.sector_scheme = 'IDS_BBG'
          AND a.is_current = TRUE
          AND a.security_durable_key IS NOT NULL
    ) b4 ON s.security_durable_key = b4.security_durable_key
    WHERE h.benchmark_durable_key IN {ids_clause}
      AND CAST(d.full_date AS DATE) IN {dates_clause}
    """

    print('*********************************************')
    print('Loading benchmark data as of', dates_clause, '...')
    print('*********************************************')
    print(query)

    dfBMKHoldings = pd.read_sql(query, lib_utils.conndremio)
    return dfBMKHoldings


# -----------------------------
# Indexes
# -----------------------------

def _load_index_holdings(
    ids: Sequence[int],
    sDate: Sequence[str],
    position_type: str = "C",
) -> pd.DataFrame:
    """
    Load index holdings from Dremio.

    Args:
        ids: One or more index_durable_key values.
        sDate: One or more 'YYYY-MM-DD' dates.
        position_type: 'C' or 'O' (current DAL uses close weights).

    Returns:
        DataFrame of index holdings.
    """
    ids_clause = _build_in_clause(ids)
    dates_clause = _build_in_clause(sDate)

    query = f"""
    SELECT
        CAST(d.full_date AS DATE) AS effective_date,
        'Index' AS portfolio_type,
        'index_durable_key' AS durable_key_type,
        b.index_durable_key AS durable_key,
        'index_key' AS key_type,
        b.index_durable_key AS portfolio_id,
        -b.index_durable_key AS benchmark_id,
        b.index_name AS portfolio_name,
        b.index_cdve_code,
        b.market_sector_des AS asset_class,
        s.country_issue_name,
        e.country_name,
        e.country_code3,
        s.currency_traded,
        'NA' AS ml_high_yield_sector,
        s.agi_instrument_asset_type_description,
        s.security_durable_key,
        s.long_name,
        s.issuer_long_name,
        s.agi_instrument_type_description,
        g.GICS AS gics_sector_description,
        bics.BICS AS bics_sector_description,
        b3.BICS AS bics_l3_description,
        b4.BBG AS BBG_l2_description,
        s.collat_type,
        'NA' AS gics_industry_group_description,
        'NA' AS gics_industry_description,
        s.bloomberg_security_typ,
        s.id_isin,
        s.id_cusip,
        s.security_des,
        s.issue_description,
        s.issuer_id_bloomberg_company AS ID_BB_COMPANY,
        h.constituent_weight_in_index / 100 AS weight
    FROM {lib_utils.dictDremio['dfIndexConst']} h
    JOIN {lib_utils.dictDremio['dfIndexDim']} b
      ON h.index_durable_key = b.index_durable_key
     AND b.is_current = TRUE
    JOIN {lib_utils.dictDremio['dfDateDim']} d
      ON d.date_key = h.date_key
    LEFT JOIN {lib_utils.dictDremio['dfSecurityMasterDim']} s
      ON s.security_durable_key = h.security_durable_key
     AND s.is_current = TRUE
    LEFT JOIN {lib_utils.dictDremio['dfCountryDim']} e
      ON s.country_issue = e.country_code
    LEFT JOIN (
        SELECT DISTINCT a.security_durable_key, l1_desc AS GICS
        FROM {lib_utils.dictDremio['dfSecuritySector']} a
        LEFT JOIN {lib_utils.dictDremio['dfSecurityMaster']} b
          ON a.security_durable_key = b.security_durable_key
        WHERE a.sector_scheme = 'IDS_GICS'
          AND a.is_current = TRUE
          AND a.security_durable_key IS NOT NULL
    ) g ON s.security_durable_key = g.security_durable_key
    LEFT JOIN (
        SELECT DISTINCT a.security_durable_key, l1_desc AS BICS
        FROM {lib_utils.dictDremio['dfSecuritySector']} a
        LEFT JOIN {lib_utils.dictDremio['dfSecurityMaster']} b
          ON a.security_durable_key = b.security_durable_key
        WHERE a.sector_scheme = 'IDS_BICS'
          AND a.is_current = TRUE
          AND a.security_durable_key IS NOT NULL
    ) bics ON s.security_durable_key = bics.security_durable_key
    LEFT JOIN (
        SELECT DISTINCT a.security_durable_key, l3_desc AS BICS
        FROM {lib_utils.dictDremio['dfSecuritySector']} a
        LEFT JOIN {lib_utils.dictDremio['dfSecurityMaster']} b
          ON a.security_durable_key = b.security_durable_key
        WHERE a.sector_scheme = 'IDS_BICS'
          AND a.is_current = TRUE
          AND a.security_durable_key IS NOT NULL
    ) b3 ON s.security_durable_key = b3.security_durable_key
    LEFT JOIN (
        SELECT DISTINCT a.security_durable_key, l2_desc AS BBG
        FROM {lib_utils.dictDremio['dfSecuritySector']} a
        LEFT JOIN {lib_utils.dictDremio['dfSecurityMaster']} b
          ON a.security_durable_key = b.security_durable_key
        WHERE a.sector_scheme = 'IDS_BBG'
          AND a.is_current = TRUE
          AND a.security_durable_key IS NOT NULL
    ) b4 ON s.security_durable_key = b4.security_durable_key
    WHERE h.index_durable_key IN {ids_clause}
      AND CAST(d.full_date AS DATE) IN {dates_clause}
    """

    print('*********************************************')
    print('Loading index holdings as of', dates_clause, '...')
    print('*********************************************')
    print(query)

    dfIdx = pd.read_sql(query, lib_utils.conndremio)
    dfIdx = pd.merge(
        dfIdx,
        lib_utils.dfHierarchy,
        how="left",
        on="ID_BB_COMPANY",
        validate="m:1",
    )
    return dfIdx


# -----------------------------
# Custom XLS portfolios
# -----------------------------

def _load_xls_portfolio(
    lDatnames: Sequence[str],
    lBenchmarkNames: Sequence[str],
    lPortfolioTypes: Sequence[str],
    lFolder: str,
    sLeftOn: str = "id_isin",
) -> pd.DataFrame:
    """
    Load custom portfolios from Excel files and enrich with security/country metadata.

    Args:
        lDatnames: Filenames (within lFolder) of the XLS sources.
        lBenchmarkNames: Display names (same length as lDatnames).
        lPortfolioTypes: Types (e.g., 'Universe', 'Benchmark') matching lDatnames.
        lFolder: Base folder path (ending with / or \\).
        sLeftOn: 'id_isin' (default) or 'ID_BB_COMPANY'/'issuer_id_bloomberg_company'.

    Returns:
        Concatenated/enriched DataFrame.
    """
    print('*********************************************')
    print('Loading customized portfolios...')
    print('*********************************************')

    dfs: List[pd.DataFrame] = []
    for datname, bench_name, ptype in zip(lDatnames, lBenchmarkNames, lPortfolioTypes):
        dfB = pd.read_excel(lFolder + datname)
        dfB["Benchmark"] = bench_name
        dfB["durable_key_type"] = "xls"
        dfB["durable_key"] = bench_name
        dfB["benchmark_id"] = datname
        dfB["portfolio_id"] = datname
        dfB["portfolio_name"] = bench_name
        dfB["portfolio_type"] = ptype
        dfs.append(dfB)

    dfAllBenchmarks = pd.concat(dfs, ignore_index=True)

    if sLeftOn == "id_isin":
        dfAllBenchmarks = pd.merge(
            dfAllBenchmarks,
            lib_utils.dfSecurities,
            how="left",
            left_on="id_isin",
            right_on="ID_ISIN",
            validate="m:1",
        )
        sec_des_col = "SECURITY_DES"
    elif sLeftOn in ("ID_BB_COMPANY", "issuer_id_bloomberg_company"):
        dfAllBenchmarks = pd.merge(
            dfAllBenchmarks,
            lib_utils.dfHierarchy,
            how="left",
            left_on=sLeftOn,
            right_on="ID_BB_COMPANY",
            validate="m:1",
        )
        sec_des_col = "LONG_COMP_NAME"
    else:
        raise ValueError("sLeftOn must be 'id_isin' or a BB company id column")

    # Cast ID columns to float as in original function
    for col in ("ID_BB_COMPANY", "ID_BB_PARENT_CO", "ID_BB_ULTIMATE_PARENT_CO"):
        if col in dfAllBenchmarks.columns:
            dfAllBenchmarks[col] = dfAllBenchmarks[col].astype(float)

    dfAllBenchmarks["issuer_long_name"] = dfAllBenchmarks.get("LONG_COMP_NAME", dfAllBenchmarks.get("issuer_long_name"))
    dfAllBenchmarks["long_name"] = dfAllBenchmarks[sec_des_col] if sec_des_col in dfAllBenchmarks.columns else dfAllBenchmarks.get("long_name")
    if "INDUSTRY_SECTOR" in dfAllBenchmarks.columns:
        dfAllBenchmarks["gics_sector_description"] = dfAllBenchmarks["INDUSTRY_SECTOR"].astype(str).str.upper()

    # Country enrichment
    if "ULT_PARENT_CNTRY_DOMICILE" in dfAllBenchmarks.columns:
        dfAllBenchmarks = pd.merge(
            dfAllBenchmarks,
            lib_utils.dfCountry,
            how="left",
            left_on="ULT_PARENT_CNTRY_DOMICILE",
            right_on="country_code",
            validate="m:1",
        )
        dfAllBenchmarks["country_issue_name"] = dfAllBenchmarks["country_name"]

    dfAllBenchmarks.dropna(subset=[sLeftOn], inplace=True)

    return dfAllBenchmarks


def calc_corp_eligible_univ(dfPortfolio: pd.DataFrame, iLookthrough: bool = False):
    """
    Classify portfolio holdings into corporate eligible universe and other categories.

    Args:
        dfPortfolio: DataFrame with security holdings, must include columns like:
            ['bloomberg_security_typ','agi_instrument_type_description',
             'agi_instrument_asset_type_description','gics_sector_description',
             'ml_high_yield_sector','id_cusip','benchmark_id']
        iLookthrough: If True, funds are treated as eligible lookthrough assets.

    Returns:
        dfPortfolio (with new classification columns),
        dfClassification (unique combinations of attributes for corporates),
        lSovTypes (list of sovereign instrument type descriptions)
    """

    # --- constants ---
    lSolidarityAssetCUSIPs = [
        'PP30KWR12','PPE9EMHU7','PPE1399U4','PP9I7DSD7',
        'PPE60EWP0','PPE83X7U8','PP9FDPE51','PPE1EGPJ3','PPE0AN3Y5'
    ]

    lSovTypes = [
        'Debts / Bonds / Fixed / Government',
        'Debts / Bonds / Variable / Government',
        'Debts / Bonds / Zero / Government',
        'Debts / MM / Fixed / Government',
        'Debts / MM / Variable / Government',
        'Debts / MM / Zero / Government',
        'Debts / MTN / Fixed / Government',
        'Debts / MTN / Variable / Government',
        'Debts / MTN / Zero / Government',
        'Debts / Municipals',
    ]

    lFundsDerivatives = [
        'Derivative','Fund','Funds','Futures','Forward','Listed Options',
        'OTC Options','Swaps','Open End Funds','Interest Rate Swap',
        'Rights','Certificates','FX Forward','Listed Derivatives',
        'Credit Default Swap','Futures -Financial','Total Return Swap',
        'Structured Other','Equity Option','Cash Option','Commodity Future Option'
    ]

    # --- normalization / overrides ---
    df = dfPortfolio.copy()

    df.loc[
        (df.bloomberg_security_typ.isin(['Unit','Stapled Security'])) &
        (df.agi_instrument_type_description == 'Funds'),
        'agi_instrument_asset_type_description'
    ] = 'Equities'

    df.loc[
        df.agi_instrument_type_description == 'Funds / REIT',
        'agi_instrument_asset_type_description'
    ] = 'Equities'

    # --- boolean flags ---
    df['Cash'] = df['agi_instrument_asset_type_description'].isin([
        'Referential Instruments','Cash','Cash Equivalent','Not Available','Account'
    ])

    df['Private Equity'] = df['bloomberg_security_typ'].eq('Private Eqty')
    df['Funds and Derivatives'] = df['agi_instrument_asset_type_description'].isin(lFundsDerivatives)
    df['Real Estate'] = df['agi_instrument_asset_type_description'].isin(['Real Estate','Mortgage'])
    df['Loan'] = df['agi_instrument_asset_type_description'].eq('Loan')
    df['sov1'] = df['ml_high_yield_sector'].isin(['QGVT','SOV'])
    df['sov2'] = df['agi_instrument_type_description'].isin(lSovTypes)
    df['sov3'] = df['agi_instrument_asset_type_description'].isin(['Government','Treas','Sovereign'])

    # gics override for sov3
    df.loc[df['gics_sector_description'].str.upper() == 'GOVERNMENT', 'sov3'] = True

    # aggregate sov
    df['Sovereigns'] = df[['sov1','sov2','sov3']].max(axis=1)
    df['Scope_Sov'] = df['Sovereigns']

    # solidarity list
    df['Solidarities'] = df['id_cusip'].isin(lSolidarityAssetCUSIPs)

    # eligible corporates = NOT in any excluded asset class
    exclusions = ['Cash','Funds and Derivatives','Sovereigns','Solidarities','Private Equity','Real Estate','Loan']
    df['Corporates eligible assets'] = ~df[exclusions].max(axis=1).astype(bool)

    # --- overrides ---
    # real estate funds can be eligible
    df.loc[
        (df.gics_sector_description.isin(['REAL ESTATE','Real Estate'])) &
        (df.agi_instrument_asset_type_description == 'Funds'),
        'Corporates eligible assets'
    ] = True

    # specific benchmarks allow sovereigns as eligible
    df.loc[
        (df.benchmark_id.isin([-344,166,94795])) &
        (df.Sovereigns == True),
        'Corporates eligible assets'
    ] = True

    # --- scopes ---
    df['Scope_Corp'] = df['Corporates eligible assets']
    df['Scope_NAV'] = True
    df['Scope_Inv'] = ~df[['Cash','Funds and Derivatives']].max(axis=1).astype(bool)

    # lookthrough funds if requested
    if iLookthrough:
        mask_funds = df['agi_instrument_asset_type_description'] == 'Funds'
        df.loc[mask_funds, ['Scope','Scope_Inv','Scope_Corp']] = True

    # --- check table for unique classification combos ---
    dfClassification = (
        df.loc[df['Corporates eligible assets'] == True,
               ['agi_instrument_asset_type_description','agi_instrument_type_description',
                'ml_high_yield_sector','bloomberg_security_typ']]
          .drop_duplicates()
    )

    return df, dfClassification, lSovTypes

def item_positions(lItems, lList):
    """Return indices of all items in lList that are in lItems."""
    return [i for i, val in enumerate(lList) if val in lItems]

def strip_tz_from_object_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Remove timezone info from datetime-like columns (object or datetime64[ns, tz])."""
    
    # Handle proper tz-aware datetime64 columns
    for col in df.select_dtypes(include=["datetimetz"]).columns:
        df[col] = df[col].dt.tz_localize(None)
    
    # Handle datetime objects stored in object dtype
    for col in df.select_dtypes(include="object"):
        if df[col].apply(lambda x: isinstance(x, datetime) and x.tzinfo is not None).any():
            df[col] = df[col].map(lambda x: x.replace(tzinfo=None) if isinstance(x, datetime) else x)
    
    return df

def excel_export(writer, exports):
    """
    Export multiple DataFrames to Excel with formatting.

    Parameters
    ----------
    writer : pd.ExcelWriter
        Open Excel writer (context handled outside).
    exports : list[dict]
        Each dict must contain:
            - data : DataFrame
            - sSheetName : str
            - sPctCols, sBars, sWide, sTrueFalse, sFalseTrue, sColsHighlight : list[str] or []
    """
    workbook = writer.book

    # Formats
    fmtpct = workbook.add_format({"num_format": "0.0%"})
    fmtdcml = workbook.add_format({"num_format": "0.0"})
    header_fmt = workbook.add_format({
        "bold": True, "text_wrap": True, "valign": "top", "border": 1
    })
    header_fmt_highlight = workbook.add_format({
        "bold": True, "text_wrap": True, "valign": "top",
        "fg_color": "#76933C", "border": 1
    })
    falsefmt = workbook.add_format({"bg_color": "#FF0000"})
    truefmt = workbook.add_format({"bg_color": "#FF0000"})

    for export in exports:
        df = strip_tz_from_object_cols(export["data"]).reset_index(drop=True)

        # Reorder columns
        a_set = set(lib_utils.lSortColumns)
        b_set = set(df.columns)
        srtcols = [*(x for x in lib_utils.lSortColumns if x in b_set),
                   *(x for x in df.columns if x not in a_set)]
        df = df.reindex(columns=srtcols)

        sheet = export["sSheetName"]
        df.to_excel(writer, sheet_name=sheet, freeze_panes=(1, 0), index=False)
        ws = writer.sheets[sheet]

        # Zoom + autofilter
        ws.set_zoom(75)
        ws.autofilter(0, 0, len(df), len(df.columns) - 1)

        # Column list
        cols = df.columns.tolist()

        # Default decimal format
        for n, col in enumerate(cols):
            if col not in lib_utils.lVarListInt:
                ws.set_column(n, n, None, fmtdcml)

        # Percent columns
        for n in item_positions(export.get("sPctCols", []), cols):
            ws.set_column(n, n, None, fmtpct)

        # Wide columns
        for n in item_positions(export.get("sWide", []), cols):
            ws.set_column(n, n, 50)

        # Data bars
        for n in item_positions(export.get("sBars", []), cols):
            ws.conditional_format(1, n, len(df), n,
                                  {"type": "data_bar", "bar_border_color": "#1F497D"})

        # Conditional TRUE/FALSE formatting
        for n in item_positions(export.get("sTrueFalse", []), cols):
            ws.conditional_format(1, n, len(df), n,
                                  {"type": "text", "criteria": "containing",
                                   "value": "FALSE", "format": falsefmt})
        for n in item_positions(export.get("sFalseTrue", []), cols):
            ws.conditional_format(1, n, len(df), n,
                                  {"type": "text", "criteria": "containing",
                                   "value": "TRUE", "format": truefmt})

        # Header formatting + comments
        for col_num, col in enumerate(cols):
            fmt = header_fmt_highlight if col in export.get("sColsHighlight", []) else header_fmt
            ws.write(0, col_num, col, fmt)
            if col in lib_utils.dictDefinitions:
                ws.write_comment(0, col_num, lib_utils.dictDefinitions[col],
                                 {"x_scale": 1.2, "y_scale": 0.8})

def calc_decimal(df, lColNames):
    """
    Divide selected columns by 100 (convert percentages to decimals).

    Parameters
    ----------
    df : pd.DataFrame
        Input DataFrame.
    lColNames : list[str]
        List of column names to transform.

    Returns
    -------
    pd.DataFrame
        Modified DataFrame with decimals.
    """
    for col in lColNames:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce") / 100
    return df

def mylambda(x, lVar):
    """
    Custom aggregator: weighted averages, total weight, and unique issuer count.
    """
    # Weighted average (ignores NaNs)
    weighted_avg = np.average(
        x[lVar].astype(float),
        weights=x["weight"],
        axis=0
    )

    # Wrap into series with suffix
    result = pd.Series(weighted_avg, index=[f"{col}_wavg" for col in lVar])

    # Add totals
    result["weight"] = x["weight"].sum()
    result["issuer"] = x["ID_BB_COMPANY"].nunique()

    return result


def calc_group_average(df, lVar, sGroupVar):
    """
    Calculate group-level weighted averages and add an 'All' category.
    """
    df[lVar] = df[lVar].astype(float)
    df[sGroupVar] = df[sGroupVar].fillna("_OTHER")

    group_vars = lib_utils.lGroupByVariablesShort + [sGroupVar]

    def _aggregate(subset):
        return mylambda(subset, lVar)

    # Regular groups
    dfRes = df.groupby(group_vars, dropna=False).apply(_aggregate)
    dfRes["all issuer"] = dfRes.groupby(level=lib_utils.lGroupByVariablesShort)["issuer"].transform("sum")
    dfRes["issuer_pct"] = dfRes["issuer"] / dfRes["all issuer"]

    # "All" groups
    dfAll = df.copy()
    dfAll[sGroupVar] = "All"
    dfAll = dfAll.groupby(group_vars, dropna=False).apply(_aggregate)
    dfAll["all issuer"] = dfAll.groupby(level=lib_utils.lGroupByVariablesShort)["issuer"].transform("sum")
    dfAll["issuer_pct"] = dfAll["issuer"] / dfAll["all issuer"]

    # Combine
    result = (
        pd.concat([dfRes, dfAll])
        .reset_index()
        .sort_values(lib_utils.lGroupByVariablesShort + ["weight"], ascending=False)
        .drop("all issuer", axis=1)
    )

    return result

def sort_cols(lVarsIn):
    """
    Reorder columns: first those in lib_utils.lSortColumns (if present),
    then all remaining in their original order.
    """
    lVarsOut = [c for c in lib_utils.lSortColumns if c in lVarsIn]
    lVarsOut += [c for c in lVarsIn if c not in lib_utils.lSortColumns]
    return lVarsOut

def concat_portfolio_benchmark(dfAllPortfolios, dfBenchmarkBridge):
    """
    Concatenate portfolio and benchmark holdings into groups.
    
    Args:
        dfAllPortfolios (pd.DataFrame): Portfolio and benchmark holdings.
        dfBenchmarkBridge (pd.DataFrame): Bridge mapping portfolio_id to benchmark_durable_key and group name.
    
    Returns:
        pd.DataFrame: Combined DataFrame with portfolio + benchmark for each group.
    """
    out_list = []

    for _, row in dfBenchmarkBridge.iterrows():
        subset = dfAllPortfolios.loc[
            dfAllPortfolios["portfolio_id"].isin([row["portfolio_id"], row["benchmark_durable_key"]])
        ].copy()
        subset["portfolio_benchmark_group_name"] = row["portfolio_benchmark_group_name"]
        out_list.append(subset)

    return pd.concat(out_list, ignore_index=True)

def calc_active_weights(dfPfBenchmark, dfBenchmarkBridge, sVar, lVarList):
    """
    Calculate active weights and contributions vs. benchmark.

    Args:
        dfPfBenchmark (pd.DataFrame): Portfolio and benchmark holdings (with 'weight', 'Scope_Corp', etc.).
        dfBenchmarkBridge (pd.DataFrame): Mapping of portfolios to benchmarks.
        sVar (str): Variable of interest (numeric column).
        lVarList (list[str]): List of grouping variables (must include portfolio/benchmark IDs & names).

    Returns:
        pd.DataFrame: Active weights/contributions with portfolio vs. benchmark diff.
    """
    # Keep only corporate scope
    df = dfPfBenchmark.loc[dfPfBenchmark["Scope_Corp"] == True].copy()

    # Contribution = weight * variable
    df["contribution"] = df["weight"] * df[sVar]

    # Normalize weights for rows with data (avoid bias from missing sVar)
    mask = ~df[sVar].isna()
    df.loc[mask, "adjweight"] = df.loc[mask].groupby(
        ["portfolio_id", "portfolio_name", "benchmark_id", "benchmark_name"]
    )["weight"].transform("sum")

    df["contribution"] = df["contribution"] / df["adjweight"]

    # Expand with benchmark groups
    df = concat_portfolio_benchmark(df, dfBenchmarkBridge)

    # Grouping (sum weights & contributions, mean of sVar)
    lVarListBench = [s for s in lVarList if s not in ["portfolio_id", "portfolio_name", "benchmark_id", "benchmark_name"]]
    df = (
        df.groupby(lVarList, dropna=False)
          .agg({
              "weight": lambda x: x.sum(min_count=1),
              "contribution": lambda x: x.sum(min_count=1),
              sVar: "mean"
          })
          .reset_index()
    )

    # Pivot Fund vs Benchmark
    df = (
        df.groupby(lVarListBench, sort=False, dropna=False)[["weight", "contribution", sVar]]
          .first()
          .unstack("portfolio_type")
    )
    df.columns = [f"{col[0]}_{col[1]}" for col in df.columns]

    # Fill sVar using fund first, then benchmark
    df[sVar] = df.get(f"{sVar}_Fund", np.nan).fillna(df.get(f"{sVar}_Benchmark", np.nan))

    # Drop helper columns
    df = df.drop(columns=[f"{sVar}_Fund", f"{sVar}_Benchmark"], errors="ignore")

    # Active contribution difference
    df["diff"] = df["contribution_Fund"].fillna(0) - df["contribution_Benchmark"].fillna(0)

    return df

def calc_period_end(period="month", year_start=2018, year_end=2025):
    """
    Load period-end dates from Dremio.

    Args:
        period (str): 'year', 'quarter', or 'month'.
        year_start (int): First year to include.
        year_end (int): Last year to include.

    Returns:
        pd.DataFrame: Period-end dates.
    """
    # Map period to SQL month filter
    if period == "year":
        month_filter = "12"
    elif period == "quarter":
        month_filter = "3,6,9,12"
    elif period == "month":
        month_filter = None
    else:
        raise ValueError("period must be 'year', 'quarter', or 'month'")

    sql = f"""
        SELECT * 
        FROM {lib_utils.dictDremio['dfDateDim']} 
        WHERE last_business_day_in_month_flag = 'Y'
          AND "year" BETWEEN {year_start} AND {year_end}
    """

    if month_filter:
        sql += f" AND month IN ({month_filter})"

    sql += " ORDER BY full_date DESC"

    return pd.read_sql(sql, lib_utils.conndremio)

def get_names(dfPortfolios, max_api=100):
    """
    Fill missing issuer names in portfolios by fetching from Bloomberg API.

    Args:
        dfPortfolios (pd.DataFrame): Portfolio dataframe with ID_BB_COMPANY and issuer_long_name.
        max_api (int): Max number of IDs allowed for Bloomberg API call (default=100).
        
    Returns:
        pd.DataFrame: Updated portfolio dataframe with missing issuer names filled.
    """
    # Find missing issuer names
    lMissings = (
        dfPortfolios.loc[dfPortfolios['issuer_long_name'].isna(), 'ID_BB_COMPANY']
        .dropna()
        .drop_duplicates()
    )

    print("**************************************")
    print(f"Looking for {len(lMissings)} missing issuer names in Bloomberg...")
    print("**************************************")

    if len(lMissings) > max_api:
        print(f"Too many IDs: {len(lMissings)} (limit {max_api})")
        if os.environ.get("BBG_API_PASSWORD") != "BBG":
            raise Exception("Not allowed to prevent API freeze")

    if len(lMissings) == 0:
        return dfPortfolios  # nothing to do

    # Prepare Bloomberg tickers
    lIDBB = ("/companyid/" + lMissings.astype(str)).to_list()

    # Fetch from Bloomberg
    lib_utils.connBBG.start()
    dfBBG = lib_utils.connBBG.ref(lIDBB, ["LONG_COMP_NAME", "COUNTRY_FULL_NAME"])

    # Extract ID_BB_COMPANY from ticker string
    dfBBG["ID_BB_COMPANY"] = dfBBG.ticker.str[11:].astype(float)

    # Reshape
    dfBBG = (
        dfBBG.pivot_table(
            index="ID_BB_COMPANY", 
            columns="field", 
            values="value", 
            aggfunc="first"
        )
        .reset_index()
    )

    # Merge back
    dfPortfolios = pd.merge(
        dfPortfolios,
        dfBBG,
        how="left",
        on="ID_BB_COMPANY",
        validate="m:1",
        suffixes=("", "_BBG"),
    )

    # Fill missing values
    dfPortfolios["issuer_long_name"] = dfPortfolios["issuer_long_name"].fillna(dfPortfolios["LONG_COMP_NAME_BBG"])
    dfPortfolios["long_name"] = dfPortfolios["long_name"].fillna(dfPortfolios["LONG_COMP_NAME_BBG"])
    dfPortfolios["country_issue_name"] = dfPortfolios["country_issue_name"].fillna(dfPortfolios["COUNTRY_FULL_NAME"])

    return dfPortfolios

import os

def search_cedar_vehicles(iRefresh, lIdentifiers=None, sType=None):
    """
    Search and filter Cedar vehicle data.

    Args:
        iRefresh (bool): If True, reload from Dremio and overwrite local pickle. 
                         If False, load cached pickle.
        lIdentifiers (list, optional): List of identifier patterns to filter on. Default = None.
        sType (str, optional): Column name to apply identifier filter on. Default = None.

    Returns:
        pd.DataFrame: Filtered vehicles dataframe.
    """
    pkl_path = os.path.join(lib_utils.sWorkFolder, "dfVehicles.pkl")

    # Refresh or load from pickle
    if iRefresh:
        sql = f"""
            SELECT *
            FROM {lib_utils.dictDremio['dfVehicleDim']}
            WHERE is_current = True
        """
        dfVehicles = pd.read_sql(sql, lib_utils.conndremio)
        dfVehicles.to_pickle(pkl_path)
    else:
        dfVehicles = pd.read_pickle(pkl_path)

    # Optional filtering
    if lIdentifiers and sType:
        # Convert to str and filter by regex OR expression
        dfVehicles = dfVehicles.loc[
            dfVehicles[sType].astype(str).str.contains("|".join(lIdentifiers), na=False)
        ]

    # Keep only active leading share classes
    dfVehicles = dfVehicles.loc[dfVehicles.vehicle_status == "Active"]
    dfVehicles = dfVehicles.loc[dfVehicles.is_leading_share_class == "Y"]

    return dfVehicles


def split_eq_fi_sleeves(dfAllPortfolios, lPortfolioIDs=None):
    """
    Splits portfolios into Equity (EQ) and Fixed Income (FI) sleeves
    and normalizes their weights within each sleeve.

    Args:
        dfAllPortfolios (pd.DataFrame): Input portfolio dataframe.
        lPortfolioIDs (list, optional): List of portfolio IDs to split.
                                        Defaults to lib_utils.EQ_FI_SPLIT_IDS.

    Returns:
        pd.DataFrame: Modified dataframe with EQ and FI sleeves.
    """
    if lPortfolioIDs is None:
        lPortfolioIDs = lib_utils.EQ_FI_SPLIT_IDS

    # Extract portfolios to split
    dfPortfolios = dfAllPortfolios.loc[dfAllPortfolios.portfolio_id.isin(lPortfolioIDs)].copy()
    # Keep the rest as-is
    dfAllPortfolios = dfAllPortfolios.loc[~dfAllPortfolios.portfolio_id.isin(lPortfolioIDs)].copy()

    # Split into FI and EQ
    dfFI = dfPortfolios.loc[dfPortfolios.agi_instrument_asset_type_description == "Debts"].copy()
    dfEQ = dfPortfolios.loc[dfPortfolios.agi_instrument_asset_type_description == "Equities"].copy()

    # Normalize weights within each sleeve
    for df in [dfFI, dfEQ]:
        df["wgt_sum"] = df.groupby("portfolio_id")["weight"].transform("sum")
        df["weight"] = df["weight"] / df["wgt_sum"]
        df.drop(columns="wgt_sum", inplace=True)

    # Rename portfolio IDs and names for clarity
    dfFI["portfolio_id"] = dfFI["portfolio_id"].astype(str) + "_FI"
    dfFI["portfolio_name"] = dfFI["portfolio_name"] + "_FI"

    dfEQ["portfolio_id"] = dfEQ["portfolio_id"].astype(str) + "_EQ"
    dfEQ["portfolio_name"] = dfEQ["portfolio_name"] + "_EQ"

    # Combine everything
    dfAllPortfolios = pd.concat([dfAllPortfolios, dfFI, dfEQ], ignore_index=True)

    return dfAllPortfolios


def get_pm_names(dfAllPortfolios, return_list=False):
    """
    Get portfolio manager names for all Fund/Portfolio-type portfolios.

    Args:
        dfAllPortfolios (pd.DataFrame): Dataframe of portfolios.
        return_list (bool, optional): If True, return a list instead of a semicolon-separated string. Default is False.

    Returns:
        str | list: Portfolio manager last names as string (joined by ;) or list.
    """
    # Get unique portfolio durable keys for funds/portfolios
    lPortfolioIDs = (
        dfAllPortfolios.loc[dfAllPortfolios.portfolio_type.isin(["Fund", "Portfolio"])]
        ["portfolio_durable_key"]
        .dropna()
        .drop_duplicates()
    )

    if lPortfolioIDs.empty:
        return [] if return_list else ""

    # Query only needed columns
    sql = f"""
        SELECT a.portfolio_durable_key,
               b.last_name,
               b.first_name
        FROM {lib_utils.dictDremio['dfPortfolioManagerBridge']} a
        LEFT JOIN {lib_utils.dictDremio['dfPortfolioManagerDim']} b
               ON a.portfolio_manager_durable_key = b.portfolio_manager_durable_key
        WHERE a.is_current = True
          AND b.is_current = True
    """
    dfPMs = pd.read_sql(sql, lib_utils.conndremio)

    # Filter only relevant portfolios
    lList = (
        dfPMs.loc[dfPMs.portfolio_durable_key.isin(lPortfolioIDs), "last_name"]
        .dropna()
        .drop_duplicates()
        .tolist()
    )

    return lList if return_list else ";".join(lList)


def sanitize_filename(name):
    """Replace invalid filename characters with underscores."""
    return re.sub(r'[\\/*?:"<>|]', "_", name)

def save_copies_with_sheetname_suffix(file_path):
    """
    Save each sheet of an Excel file as a separate file
    with the sheet name appended to the base filename.
    
    Args:
        file_path (str): Path to the original Excel file.
    
    Returns:
        list[str]: List of saved file paths.
    """
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    sheet_names = workbook.sheetnames

    # List to store saved file paths
    saved_files = []

    # Get directory and base file name
    directory, base_name = os.path.split(file_path)
    base_name_no_ext = os.path.splitext(base_name)[0]

    for sheet_name in sheet_names:
        # Create a new workbook
        new_workbook = Workbook()
        new_workbook.remove(new_workbook.active)

        # Sanitize sheet name for file naming
        safe_sheet_name = sanitize_filename(sheet_name)

        # Copy sheet content with styles and formatting
        source = workbook[sheet_name]
        target = new_workbook.create_sheet(sheet_name)

        for row in source.iter_rows():
            for cell in row:
                new_cell = target[cell.coordinate]
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell._style = cell._style

        # Copy column widths
        for col_letter, col_dim in source.column_dimensions.items():
            target.column_dimensions[col_letter].width = col_dim.width

        # Copy row heights
        for row_idx, row_dim in source.row_dimensions.items():
            target.row_dimensions[row_idx].height = row_dim.height

        # Copy merged cells
        for merged_range in source.merged_cells.ranges:
            target.merge_cells(str(merged_range))

        # Construct new file name
        new_file_name = f"{base_name_no_ext}_{safe_sheet_name}.xlsx"
        new_file_path = os.path.join(directory, new_file_name)

        # Save the new workbook
        new_workbook.save(new_file_path)
        print(f"Saved {new_file_path}")

        saved_files.append(new_file_path)

    workbook.close()
    return saved_files
    
def compare_two_lists(dfPrev, dfCurrent, lIDs, name_col="LONG_COMP_NAME"):
    """
    Compare two DataFrames and identify NEW and REMOVED rows based on IDs.

    Args:
        dfPrev (pd.DataFrame): Previous snapshot of the dataset.
        dfCurrent (pd.DataFrame): Current snapshot of the dataset.
        lIDs (list[str]): List of column names to use as keys for comparison.
        name_col (str): Optional descriptive column to include in exits (default: 'LONG_COMP_NAME').

    Returns:
        tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
            - dfCurrentFlagged: Current DataFrame with NEW_FLAG column
            - dfAdditions: Subset of new rows
            - dfExits: Subset of removed rows (with IDs + name_col)
    """
    # --- Additions (what’s new in Current) ---
    dfCurrentFlagged = dfCurrent.merge(
        dfPrev[lIDs], how="left", on=lIDs, indicator=True
    ).rename(columns={"_merge": "NEW_FLAG"})

    dfCurrentFlagged["NEW_FLAG"] = dfCurrentFlagged["NEW_FLAG"].map({
        "left_only": "NEW",
        "right_only": "REMOVED",  # should never happen here
        "both": ""
    })

    # --- Exits (what’s missing compared to Prev) ---
    dfExits = dfPrev[lIDs + [name_col]].merge(
        dfCurrent.drop(columns=["NEW_FLAG"], errors="ignore"),
        how="left", on=lIDs, indicator=True
    ).rename(columns={"_merge": "NEW_FLAG"})

    dfExits["NEW_FLAG"] = dfExits["NEW_FLAG"].map({
        "left_only": "REMOVED",
        "right_only": "NEW",  # should never happen here
        "both": ""
    })

    # Keep only actual removals
    dfExits = dfExits[dfExits["NEW_FLAG"] == "REMOVED"][lIDs + [name_col]]

    # --- Subset of new rows ---
    dfAdditions = dfCurrentFlagged[dfCurrentFlagged["NEW_FLAG"] == "NEW"]

    return dfCurrentFlagged.copy(), dfAdditions, dfExits

def extract_and_join(input_string: str) -> str | None:
    """Extract substrings inside < > and join them with ';'."""
    if pd.isna(input_string) or not isinstance(input_string, str):
        return None

    matches = re.findall(r'<(.*?)>', input_string)
    matches = [m.strip() for m in matches if m.strip()]
    return ';'.join(matches) if matches else None

def compare_lists(list_a: str, list_b: str) -> str | None:
    """Compare two semicolon-separated lists and return items only in list_a."""
    if pd.isna(list_a) or not isinstance(list_a, str):
        return None
    if pd.isna(list_b) or not isinstance(list_b, str):
        return list_a  # if list_b missing, everything in list_a is different

    list_a_split = [x.strip() for x in list_a.split(';') if x.strip()]
    list_b_split = [x.strip() for x in list_b.split(';') if x.strip()]
    
    difference = set(list_a_split) - set(list_b_split)
    return ';'.join(sorted(difference)) if difference else None
def calculate_ratios(df: pd.DataFrame, group_col: str, id_col: str, bool_columns: list[str]) -> pd.DataFrame:
    """
    Calculate ratios of unique IDs where boolean columns are True within each group.

    Args:
        df: Input dataframe.
        group_col: Column to group by.
        id_col: Identifier column for uniqueness.
        bool_columns: List of boolean columns to compute ratios on.

    Returns:
        DataFrame with one row per group and ratio columns added.
    """
    def ratio_fn(group):
        total_ids = group[id_col].nunique()
        if total_ids == 0:
            return {f"{col}_ratio": 0.0 for col in bool_columns}
        return {
            f"{col}_ratio": group.loc[group[col], id_col].nunique() / total_ids
            for col in bool_columns
        }

    result = df.groupby(group_col).apply(ratio_fn).reset_index(name="ratios")

    # Expand the dictionary column into separate columns
    ratios_df = pd.DataFrame(result["ratios"].tolist(), index=result.index)
    result = pd.concat([result[group_col], ratios_df], axis=1)

    return result

def extract_max_year(expression):
    """
    Extract the maximum year from a string containing 'FYxxxx' or 'xxxx'.

    Args:
        expression: str, numeric, or NaN.

    Returns:
        float: maximum year found, or np.nan if none.
    """
    if pd.isna(expression):
        return np.nan

    expression = str(expression)

    # Find all 4-digit year numbers, with or without 'FY' prefix
    matches = re.findall(r'(?:FY)?(\d{4})', expression)

    if not matches:
        return np.nan

    years = [float(y) for y in matches]
    return max(years)

# Columns shared across all holding loaders
def _ensure_columns(df: pd.DataFrame, columns: Sequence[str]) -> pd.DataFrame:
    """Ensure all required columns exist in ``df``."""

    for col in columns:
        if col not in df.columns:
            df[col] = np.nan
    return df[list(columns)]


def _standardize_holdings(df: pd.DataFrame, columns: Sequence[str]) -> pd.DataFrame:
    """Return ``df`` with required columns, types, and default values."""

    df = _ensure_columns(df, columns)
    df = df.fillna({col: lib_utils.HOLDING_DEFAULTS[col] for col in columns})
    df = df.astype({col: lib_utils.HOLDING_DTYPES[col] for col in columns}, errors="ignore")

    df = _assign_region(df)
    return df


def _assign_region(df: pd.DataFrame) -> pd.DataFrame:
    df["region"] = np.nan
    df.loc[df.country_code3.isin(lib_utils.lNA), "region"] = "NA"
    df.loc[df.country_code3.isin(lib_utils.lAPAC), "region"] = "APAC"
    df.loc[df.country_code3.isin(lib_utils.lEROP), "region"] = "EROP"
    df["region"] = df["region"].fillna("EM")
    return df


def load_holdings(source: str, *args, columns: Sequence[str] = lib_utils.HOLDING_COLUMNS, **kwargs):
    try:
        loader = lib_utils.HOLDINGS_LOADERS[source.lower()]
    except KeyError as exc:
        raise ValueError(f"Unknown data source: {source!r}") from exc

    result = loader(*args, **kwargs)
    if isinstance(result, tuple):
        df = _standardize_holdings(result[0], columns)
        return (df,) + result[1:]
    return _standardize_holdings(result, columns)
